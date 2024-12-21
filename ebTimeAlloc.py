#!/usr/bin/env python
# coding: utf-8
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font

import uml_V2.uml_lib.ebAPI_lib as ebAPI

#import newebapi_proj as project_data
#import newebapi_budg as budget_data


"""
230123
- Rebuilt/Modified the code to support the data from New(REST) API

211103
- Eliminating special team roles, except for FF&E
- Filtering out Planners/Planning lead
210304
Filter out Planners/Planning lead


210120
- Using Budget Report -> UML - PS Project ID Report, save as CSV

201229 Rewriting with API to Budget Request project, eliminating CSV file/report - we
can get the PS project IDs from API Process_BDFRQ

201020 Add project name as col/field 2

201006 Adding ability to produce separate file for each PM/Planner
    APPROACH: Loop once to get all info. Collect list of participants

Custom_Total_Current_Working_EstimateTPC

Project Manager
FMP Number
Projects (this would be projects by PM per line)
Time Period being reported on (I can add the time period if needed)
TPC
PS Project Number (this is a project cost field and note there could be multiple so we would need all of them) You can find it by going to Cost/Funding Sources then clicking on the custom field icon.  Can you see if you can pull this info?
"""



TeamRoles = ["Project Manager","Assistant Project Manager","FF&E Team Member"]

def read_EBspeedtype_report(theCSV):
    FMP_PSproj = {}

    with open(theCSV) as csvfile:
        STdata = csv.DictReader(csvfile,delimiter=',') #quotecharacter?
        for r in STdata:
            if r["FMP Number"] not in FMP_PSproj:
                # New FMP number, create new dictionary entry with list
                FMP_PSproj[r["FMP Number"]] = r["Project ID"]
            else:
                FMP_PSproj[r["FMP Number"]] += "," + r["Project ID"]

    #for f in FMP_PSproj:
        #print "From read report", f, FMP_PSproj[f]
    return FMP_PSproj

def create_Excel():
    wbOut = Workbook()
    #sheetName = "EBimport"
    #outWS = wbOut.create_sheet(sheetName,0)
    # use default, "Sheet" - CONSIDER renaming to "EBimport"
    outWS = wbOut.active
    return(wbOut,outWS)

## Importing this directly from the eb.ebApI_lib.py - Redundant, Hence removed.
#
# def get_Projects_allData():
#     #theFields = ["Custom_Campus1","Custom_Building","Custom_Planned_Occupant","Custom_FM_Department_Lead","Custom_Project_Planner","Custom_Project_Manager"]
#     jObj = ebAPI.APIconnect("Projects")
#
#     return(jObj)
#

def hyphenator(s, mode):
    retVal = ""
    for c in s:
        if mode == "ADD" and c == " ":
            retVal += "_"
        elif mode =="REMOVE" and c == "_":
            retVal += " "
        else:
            retVal += c
    return retVal

def get_numeric_monthYear():
    now = datetime.now()
    #moStr = str(now.month-4)
    moStr = str(now.month+1)
    if len(moStr) < 2:
        moStr = "0" + moStr
    yrStr = str(now.year)[2:]
    #print moStr, yrStr
    return (yrStr+ moStr + "_")

# ## Importing this directly from the eb.ebApI_lib.py - Redundant, Hence removed.
#def get_Budgets_allData():
#    jObj = ebAPI.APIconnect("Budgets")
#    return jObj

def get_TPC(budget_data, currProjID):
    currTPC = -1
    a = len(budget_data)

    for b in range(0,a):
        if budget_data[b]["projectId"] == currProjID:
            #print ">>>>", budget_data[b]["ProjectID"], currProjID
            #print "!!!!!!",budget_data[b]["Custom_Total_Current_Working_EstimateTPC"]
            try:
                currTPC = budget_data[b]["Total Current Working Estimate/TPC"]
            except:
                currTPC = -1
            break
    if currTPC == -1:
        currTPC = None
    #print ("RETURNING ............................................... TPC", currTPC)
    return currTPC

def writeCell(currWS, r,c,val):
    outCell = currWS.cell(row=r, column=c)
    outCell.value = val

def get_Role(person, projData):
    global TeamRoles
    retVal = "None: check EB"

    roles = ""
    # what about multiple?
    for t in TeamRoles:
        if projData[t] == person:
            roles += t + ","
    roles = roles[:-1]
    return roles

def write_Headers(outXL_WS):
    # Changing headers to row 2, adding user name to row 1

    xlHeaders = {1:"#",2:"Project Manager", 3:"FMP Number", 4:"Projects",5:"Percent of Time", 6:"Time Period bring reported on", 7:"Total Current Working Estimate/TPC", 8:"PS Project Number", 9:"Role"}
    ft = Font(color="FFFFFF", bold=True)
    r = 1
    for c in range(1,5):
        outCell = outXL_WS.cell(row=r, column=c)
        outCell.style = 'Accent1'
        outCell.font = ft
    for h in xlHeaders:
        outCell = outXL_WS.cell(row=1, column=h)
        outCell.value = xlHeaders[h]
        outCell.style = 'Accent1'
        outCell.font = ft

def get_ProjData(ebProjID, ebProjData):
    #print(">>>",ebProjID)
    i = len(ebProjData)
    retVal = "PROJECT NOT FOUND"
    # Now, loop through eB Object, build rows with desired data plucked
    for j in range(0,i):
        #print(ebProjID,"<>",ebProjData[j]["projectId"] )
        if (ebProjData[j]["projectId"] == ebProjID):
            #print( "Found it!", ebProjData[j]["name"])
            retVal = ebProjData[j]
    print(retVal)
    return retVal

def main(iFile):
    # iFile should contain the data file of the input
    # oDir should contain the path to where the output of Time Allocations has to be saved
    # Modify accordingly....

    #iFile = 'DataFiles\\UML  Time Allocation includes all PMs - 202301061224.csv' - variable defined as constructor to main()
    print("Last update: August 31, 2023\nOK from ebTimeAlloc\n", iFile)
    oDir = 'B:\\xfer\\'
    #oDir = "/Users/kysgattu/FIS/BDrive/xfer/"

    global TeamRoles
    #uname = getpass.getuser()

    # 1) Connect to eb API to get Projects and Budgets data
    # ebProjData = ebAPI.get_active_project_all_data()
    ebProjData = ebAPI.getDataFromCache("ActiveProjects")
    
    #print (ebProjData[10]["FMP Number"])
    print('Active Projects data imported')
    # ebBudgetData = ebAPI.get_budget_all_data()
    ebBudgetData = ebAPI.getDataFromCache("Budgets")
    #for k in ebBudgetData[0]:
        #print(k,ebBudgetData[10][k])
    print('Budgets data imported')

    #ebProjData = project_data.get_project_all_data()
    #ebBudgetData = budget_data.get_budget_all_data()

    k = len(ebBudgetData)

    # ebFMP_PS = get_PSprojNums()
    ebFMP_PS = read_EBspeedtype_report(iFile)

    # 3) Define team member fields
    theProjs = {}
    thePeople = {}  # track all people so we can loop and produce report for each

    # 4) loop #1: through budget data. If project budget exceeds threshold, collect data
    r = 2

    # loop on all budget items
    for l in range(0, k):
        currProjData = {}
        if ebBudgetData[l]["Total Current Working Estimate/TPC"] != None:
            # Data from New API is adding commas in the budget value, so removing them for clean typecasting to float
            currTPC = float(ebBudgetData[l]["Total Current Working Estimate/TPC"].replace(",", ""))
            # August 30, 2023: check TPC threshold first, before getting project data
            if (currTPC > 1000000.):
                
                # print ("WTF?",currTPC, ebBudgetData[l]["projectName"],ebBudgetData[l]["status"])
                # Project may have TPC > 1,000,000 but not be Active or TD Active
                
                currProjData = get_ProjData(ebBudgetData[l]["projectId"],ebProjData)
                if currProjData == "PROJECT NOT FOUND":
                    print("Project not active: skip",ebBudgetData[l]["projectName"])
                else:
                    #print (currTPC, currProjData["name"])
                    # if (currTPC > 1.):
                    # Use TEST project, APP200, for debugging/development
                    # if currProjData["Status"] == "Active" or currProjData["Status"] == "TD Active" or currProjData["Custom_Project_Number"] == "APP200" :
                    if currProjData["status"] == "Active" or currProjData["status"] == "TD Active":
                        # print "TPC above 1,000,000: do it", currTPC
                        # print(currProjData['FMP Number'])
                        currFMP = currProjData["FMP Number"]
                        # print(currFMP)
                        theProjs[currFMP] = currProjData
                        # print(currProjData)
                        for t in TeamRoles:
                            if currFMP == "APP200":
                                print("???", t, "\n")
                            try:
                                # print ("ok")
                                if theProjs[currFMP][t] != None:
                                    # print (t, theProjs[currFMP][t])
                                    # need to know if the fmp exists in person's list
                                    if theProjs[currFMP][t] in thePeople:
                                        # print ("Already in the list")
                                        if currFMP not in thePeople[theProjs[currFMP][t]]:
                                            thePeople[theProjs[currFMP][t]].append(currFMP)
                                    else:
                                        # print ("got here")
                                        thePeople[theProjs[currFMP][t]] = [currFMP]
                                else:
                                    # print (">nobody in role", t)
                                    x = 1
                            except:
                                # print "nobody in role", t
                                x = 1
                        # print theProjs
            else:
                # print "TPC below 1,000,000: ignore"
                else_filler = 1

    # 5) Now, we have project data and a list of all people from all roles. Loop on people, create an excel for each, and output
    #    a row for each project their on, listing the role(s) they're in (can they be in more than one?)
    # print thePeople

    for p in thePeople:
        print(p, thePeople[p])
        pName = hyphenator(p, "ADD")
        # print "\nOutput to Excel, for", p, pName
        currWB, currWS = create_Excel()
        ft = Font(color="FFFFFF", bold=True)
        # writeCell(currWS,1,1,("Current Projects for "+ p))

        currWS.column_dimensions['B'].width = 25.
        currWS.column_dimensions['C'].width = 8.
        currWS.column_dimensions['D'].width = 45.
        currWS.column_dimensions['G'].width = 15.
        currWS.column_dimensions['H'].width = 35.

        write_Headers(currWS)
        monthStr = get_numeric_monthYear()
        ofile = oDir + monthStr + "_TimeAlloc_" + pName + ".xlsx"
        r = 2
        for fmp in thePeople[p]:
            currRole = get_Role(p, theProjs[fmp])
            pRole = hyphenator(currRole[7:], "REMOVE")
            # {1:"Index",2:"Project Manager", 3:"FMP", 4:"Project Name",5:"Percent of Time", 6:"Time Period", 7:"TPC", 8:"PS PO Numbers", 9:"Role"}
            # print "\t", fmp, ",", pRole # That strips off the "Custom_". Should replace _ with space
            writeCell(currWS, r, 1, str(r - 2))
            writeCell(currWS, r, 2, p)
            writeCell(currWS, r, 3, fmp)
            writeCell(currWS, r, 4, theProjs[fmp]["name"])
            try:
                writeCell(currWS, r, 8, ebFMP_PS[fmp])
            except:
                writeCell(currWS, r, 8, "None listed")
            writeCell(currWS, r, 9, pRole)
            # print((get_TPC(theProjs[fmp]["projectId"], ebBudgetData)))
            theTPC = float(get_TPC(ebBudgetData, theProjs[fmp]["projectId"]).replace(",", ""))
            # print theTPC
            outCell = currWS.cell(row=r, column=7)
            outCell.value = theTPC
            outCell.number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
            r += 1  # row we're writing to
        currWB.save(ofile)
    return oDir


if __name__ == "__main__":
    #main('DataFiles\\UML  Time Allocation includes all PMs - 202301061224.csv')
    #main('/Users/kysgattu/FIS/BDrive/xfer/UML  Time Allocation CURRENT.csv')
    main('B:\\xfer\\UML  Time Allocation CURRENT.csv')
    #print("Select CSV")




#get_ipython().system('jupyter nbconvert --to script ebTimeAlloc_rough.ipynb')
