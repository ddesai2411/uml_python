#!/usr/bin/env python
# coding: utf-8
import xmltodict, glob
from uml_V2.uml_lib import ebAPI_lib as eb
from uml_V2.uml_lib import web_lib as UMLweb
from uml_V2.uml_lib import ebCostLib as ebCost
#import ebAPI_lib_v2 as eb
from datetime import datetime
from openpyxl import Workbook
import os
import shutil
import pandas as pd

# 240106 Looks like we're not checking if the invoice is already in EB
# TO ADD: get "DueDate" from Buyways and add as EB "Scheduled Paid Date"

def checkUMLPAYAP(invFiles, theDir):
    PAYAP_FilesDir = theDir + 'PAYAP_Files'
    Invoice_dict = {'PAYAP Files': [], 'Non PAYAP Files': []}
    for f in invFiles:
        # print(f)
        PAYAP_flag = f.find("PAYAP")
        if PAYAP_flag != -1:
            Invoice_dict['PAYAP Files'].append(f)
        else:
            Invoice_dict['Non PAYAP Files'].append(f)
            # isPAYAP +=1
    return Invoice_dict


def checkUMLPO(currPONumber,f):
    count = 0
    # this is redundant, if we run Filter first and we're looking in \2process
    if currPONumber[0]!= 'L':
        try:
            print(f)
            # Delete the file
            os.remove(f)
            count +=1
        except:
            #print("File doesn't exist or deleted in previous invoice line")
            ignoreThis = 1
    return count

def moveProcessedFiles(source_Dir, processed_Dir):
    InvoiceFiles = glob.glob(source_Dir + "*_Invoice_*.xml")
    print ("Moving files")
    fileCount = 0
    for i in range(len(InvoiceFiles)):
        InvoiceFile = InvoiceFiles[i]
        InvoiceFileName = InvoiceFiles[i].split('\\')[-1]
        print(InvoiceFile)
        shutil.move(InvoiceFile, (processed_Dir + InvoiceFileName))
        #move_dest = os.path.join(processed_Dir, InvoiceFileName)
        # print(move_dest)
        """
        try:
            if not os.path.exists(move_dest):
                # print('e')
                shutil.move(InvoiceFile, processed_Dir)
            else:
                #print("Not moving, already moved??")
                ignoreThis = 1
                
        except Exception as e:
            print(str(e))  # Print the exception message
        """
        fileCount += 1
        if fileCount % 100 == 0:
            print(".")

def get_XMLfund(theData):
    for f in theData:
        if f["@name"] == "Fund":
            #print("GOT IT")
            if type(f["CustomFieldValue"]) == dict:
                currFund = f["CustomFieldValue"]["Value"][:-2]
                #print(currFund)
                return currFund
            else:
                multiFund = []
                for i in range(len(f["CustomFieldValue"])):
                    #print(f["CustomFieldValue"][i]["Value"])
                    currFund = f["CustomFieldValue"][i]["Value"][:-2]
                    if currFund not in multiFund:
                        multiFund.append(currFund)
                    retVal = multiFund
                PSprojects = ",".join(multiFund)
                return PSprojects

def get_Speedtypes(theData):
    for f in theData:
        if f["@name"] == "Speedtype":
            if type(f["CustomFieldValue"]) == dict:
                currST = f["CustomFieldValue"]["Value"][:-2]
                #print(currST)
                return currST
            else:
                multiST = []
                for i in range(len(f["CustomFieldValue"])):
                    #print(f["CustomFieldValue"][i]["Value"])
                    currST = f["CustomFieldValue"][i]["Value"][:-2]
                    if currST not in multiST:
                        multiST.append(currST)
                    retVal = multiST
                speedtypes = ",".join(multiST)
                #for i in range(len(multiST)):
                #speedtypes = speedtypes + "," + multiST[i]
                return speedtypes

def get_customField(theData,theField):
    retVal = "NOT FOUND"

    for f in theData:
        #print "???", f["@name"]
        if f["@name"] == theField:
            try:
                retVal =  f["CustomFieldValue"]["Value"]
            except:
                retVal = "MULTIPLE"
                #for g in f:
                #    print( "\tget_customField", g)
    #print(retVal)
    return retVal

def checkFMP(theFMP):
    retVal = True
    for c in theFMP:
        if c not in "0123456789":
            retVal = False
            break
    return retVal

def get_FMPfromPSproject(psProj):
    retVal = "FMP NOT FOUND"
    #print (">>>>", psProj)

    iFMP = psProj.find("FMP")
    iDash = psProj.find("-")
    if iFMP == -1:
        retVal = "FMP STRING NOT FOUND"
    elif iDash == -1:
        retVal = "FMP: DASH NOT FOUND"
    else:
        #print( "get_FMP >>>>", psProj[(iFMP+3):iDash])
        currFMP = psProj[(iFMP+3):iDash]
        if currFMP[0] == "0":
            currFMP = currFMP[1:]
            #print ">>>>", currFMP
        if len(currFMP) != 6:
            retVal = "FMP: LENGTH OF STRING IS WRONG"
        elif checkFMP(currFMP):
            retVal = currFMP
        else:
            retVal = "FMP: NON-NUMERIC CHARACTER"
    return retVal

def getFieldFromSource(source,field):
    retval = ''
    if source != 'Field Not Found in XML' and field in source:
        #print(type(source),type(field))
        #retval = retval + source[field]
        retval = source[field]
        return retval
    else:
        retval = retval + 'Field Not Found in XML'
        return retval

def create_Excel():
    wbOut = Workbook()
    outWS = wbOut.active
    return(wbOut,outWS)

def writeCell(currWS, r,c,val):
    outCell = currWS.cell(row=r, column=c)
    outCell.value = val

def write_Headers(outXL_WS, xlHeaders):
    for h in xlHeaders:
        outCell = outXL_WS.cell(row=1, column=h)
        outCell.value = xlHeaders[h]

def dateStrToDate(currCommitDate):
    retval = ''
    yrStr = currCommitDate[0:4]
    moStr = currCommitDate[5:7]
    dtStr = currCommitDate[8:10]
    retval = moStr + '/' + dtStr + '/' + yrStr
    return retval

def getCurrCompName(datasource):
    supplier = getFieldFromSource(datasource, 'Supplier')
    supplier_name = getFieldFromSource(supplier, 'Name')
    return supplier_name


def build_commitTypes(activePOs):
    global vendorTypes

    for p in activePOs:
        currVendor = activePOs[p]["Vendor"]  # .encode("utf-8")
        currType = activePOs[p]["CommitmentType"]  # .encode("utf-8")

        if currVendor not in vendorTypes:
            vendorTypes[currVendor] = [currType]
        else:
            if currType not in vendorTypes[currVendor]:
                vendorTypes[currVendor].append(currType)
    vendorTypes["DAEDALUS PROJECTS INC"] = ["4. Purchase Order-House Doctors & OPM's"]
    vendorTypes["DIAMOND RELOCATION INC"] = ["3. Purchase Order-Master Service Vendor/Contractor"]
    vendorTypes["EXPRESS SIGN & GRAPHICS INC"] = ["3. Purchase Order-Master Service Vendor/Contractor"]
    vendorTypes["MILLER DYER SPEARS INC"] = ["4. Purchase Order-House Doctors & OPM's"]
    vendorTypes["MODULEX NEW ENGLAND"] = ["3. Purchase Order-Master Service Vendor/Contractor"]
    vendorTypes["R-SQUARED OFFICE PANELS & FURNITURE INC"] = ["3. Purchase Order-Master Service Vendor/Contractor"]
    vendorTypes["STATE ELECTRIC CORPORATION"] = ["3. Purchase Order-Master Service Vendor/Contractor"]
    vendorTypes["TRC ENVIRONMENTAL CORPORATION"] = ["1. Purchase Order"]
    vendorTypes["TRIUMVIRATE ENVIRONMENTAL INC"] = ["1. Purchase Order"]
    vendorTypes["VAV INTL INC"] = ["4. Purchase Order-House Doctors & OPM's"]
    vendorTypes["WILLIAM LOWE & SONS CORP"] = ["3. Purchase Order-Master Service Vendor/Contractor"]

def getCommitTypeFromVendor(vendorName):
    currVendor = vendorName.upper()
    # global vendorTypes
    retVal = "NOT FOUND"
    # print type(currVendor)
    if currVendor in vendorTypes:
        # print "Found!", currVendor
        # print "+++++++++++++++++++++++++++++++++", vendorTypes[currVendor]
        try:
            if len(vendorTypes[currVendor]) == 1:
                retVal = vendorTypes[currVendor][0]
            else:
                retVal = "CHECK: MULTIPLE VALUES"
        except:
            ignoreThis = 1
            # print ".......failed on vendorTypes"
    """
    try:
        retVal = vendorTypes[v]
        if len(vendorTypes[v]) > 1:
            retVal = "CHECK TYPE: MULTIPLE VALUES"
    except:
        print ">?>?", currVendor
        retVal = "VENDOR NOT FOUND"
    """
    # print "from getCommitTypeFromVendor:\n", retVal
    return retVal

def getCurrCommitType(currPONumber,currVendor):
    if currPONumber in activePOs:
        retval = activePOs[currPONumber]['CommitmentType']
    else:
        retval = getCommitTypeFromVendor(currVendor)
    return retval

# 1. FMP Number
def getFMPNumber(activePOs,currPONumber):
    currFMP = activePOs[str(currPONumber)]['FMP'].decode()
    return currFMP

# 2. Invoice Number
def getCurrInvNumber(datasource):
    invNum = getFieldFromSource(datasource, 'InvoiceNumber')
    return invNum

# 3. Vendor Invoice #
def getCurrVendorInvNumber(datasource):
    try:
        venInvNum = getFieldFromSource(datasource, 'SupplierInvoiceNo')#.encode('ASCII')
    except:
        venInvNum = "SEE BUYWAYS"
    return venInvNum

# 4. Commitment Number
def getCurrPONum(datasource):
    currOriginCode = get_customField(datasource["CustomFieldValueSet"],"Origin Code")
    currExtReq = get_customField(datasource["CustomFieldValueSet"],"External Req #")
    #if currOriginCode == "LEB":
        #currPONum = currExtReq[:2]
    #else:
        #currPONum =  getFieldFromSource(datasource,'PONumber')
    currPONum =  getFieldFromSource(datasource,'PONumber')
    return currPONum

# 5. Description
def getCurrDescription(datasource):
    item = getFieldFromSource(datasource, 'Item')
    description = getFieldFromSource(item, 'Description')
    return description

def multilineDescription(currDesc,currLineDescList):
    #curr_list = ['']
    curr_desc = ''
    if currDesc in currLineDescList:
        #print('check')
        curr_desc = ''.join(currDesc)
        currLineDescList.append(currDesc)
        return curr_desc
    else:
        #print('here')
        currLineDescList.append(currDesc)
        #print(currLineDescList)
        curr_desc = ' | '.join(str(item) for item in currLineDescList)
        return curr_desc

def checkIfDescIsLong(currDescription):
    if len(currDescription) > 365:
        retval = "SEE BUYWAYS FOR FULL DESCRIPTIONS - " + currDescription
    else:
        retval = currDescription
    return retval

# 6. Invoice Item Amount
def getCurrItemAmnt(datasource):
    itemAmnt = float(getFieldFromSource(datasource, 'UnitPrice')) + float(getFieldFromSource(datasource, 'ShippingCharge'))
    retval = str(itemAmnt)
    return retval

# 7. Commitment Item Number
def getCurrCommitItemNum(datasource):
    currPOItemNum =  getFieldFromSource(datasource,'POLineNumber')
    return currPOItemNum

# 8. Funding Rule

def getFundingRule(currST, fundingRule):
    #print(currST)
    retval = ''
    flag = currST.find(',')
    if flag == -1:
        if currST in fundingRule.keys():
            retval = retval + fundingRule[currST]['Name']
        else:
            retval = retval + "Funding Rule Not Found"
        return retval
    else:
        fundrules = []
        speedtypes = currST.split(',')
        #print(speedtypes)
        for i in speedtypes:
            #print(i)
            if i in fundingRule.keys():
                #fundrules = ','.join(fundingRule[i]['Name'])
                fundrules.append(fundingRule[i]['Name'])
        #print(fundrules)
        if len(fundrules) != 0:
            for i in range(len(fundrules)):
                retval = ",".join(fundrules)
        else:
            retval = "Funding Rule Not Found"
        #print(retval)
        return retval

# 9. Voucher ID
def getCurrVoucherID(datasource):
    voucherID = getFieldFromSource(datasource, 'InvoiceNumber')
    return voucherID

# 10. Status
def getcurrInvStatus(datasource):
    invStatus = getFieldFromSource(datasource, '@status')
    if invStatus == "Payable":
        retval = "Paid"      # THIS WAS THE PROBLEM 240204
    elif invStatus == "Paid":
        retval = "Paid"
    else:
        retval = (invStatus + " ERROR!!!!")
    return retval

# 11. Approved Date
def getInvCurrApprovedDate(datasource):
    dateStr = getFieldFromSource(datasource, 'InvoiceCreateDate')
    date = dateStrToDate(dateStr)
    return dateStr

# 12. Amount this Period
def getCurrAmntThisPeriod(datasource):
    amount = float(getFieldFromSource(datasource, 'UnitPrice')) + float(getFieldFromSource(datasource, 'ShippingCharge'))
    retval = str(amount)
    return retval

# 13. Stored Materials
def getCurrStoredMaterials():
    return "0" # Hardcoded in old API too

# 14. Quantity This Period
def getCurrQuantityThisPeriod(datasource):
    quantity = getFieldFromSource(datasource, 'Quantity')
    return quantity


def InvXMLtoExcel(Invoice_dict):
    currFMPs = []
    currInvNums = []
    currVendorInvoiceNums = []
    currPONums = []
    currDescriptions = []
    currInvItemAmnts = []
    currCommitItemNums = []
    currFundRules = []
    currVoucherIds = []
    currStatuses = []
    currApprovedDates = []
    currAmntThisPeriods = []
    currStoredMaterials = []
    currQuantityThisPeriods = []
    PAYAPcurrFMPs = []
    PAYAPcurrInvNums = []
    PAYAPcurrPONums = []
    count = 0
    NonUMLInvs = 0
    UMLInvs = 0
    nonEB = 0
    EB = 0
    alreadyIn = 0
    notAlreadyIn = 0
    activePO = 0
    notActivePO = 0
    cancelledInvoices = 0
    EBCost = 0
    firstCost = True
    EBProcess = 0
    EBcostType20 = 0
    InProcess = 0
    Payable = 0
    AlreadyInEB = 0
    Other = 0


    PAYAPFiles = Invoice_dict['PAYAP Files']
    NonPAYAPFiles = Invoice_dict['Non PAYAP Files']
    PAYAPs = len(PAYAPFiles)
    NonPAYAPs = len(NonPAYAPFiles)

    count = 0
    tester = 0 # make this 0 for production: it limits how many files we check
    r = 2 #remove this to create seperate files for each XML
    print("ready to loop oninvoice XML", tester, count)
    for f in NonPAYAPFiles:
        if tester > 0 and count > tester:
            break
        #r = 2 # Add this to create seperate files for each XML
        # print(f)
        theFile = open(f,encoding='utf-8')
        print("Working on ", f)
        xml_content = theFile.read()
        theFile.close()
        bwdata = xmltodict.parse(xml_content,encoding='utf-8')
        headerData = bwdata['BuyerInvoiceExportMessage']['Invoice']['InvoiceHeader']
        if 'InvoiceLine' not in bwdata['BuyerInvoiceExportMessage']['Invoice']:
            print("*****Invoice Status Cancelled in ", f)
            cancelledInvoices +=1
            invoiceType = "Cancelled"

        else:
            Invlinedata = bwdata['BuyerInvoiceExportMessage']['Invoice']['InvoiceLine']
            #print(type(InvoiceLinedata))
            #print(headerData["InvoiceNumber"])

            if type(Invlinedata) == dict:
                currInvNum = getCurrInvNumber(headerData)
                currPONumber = getCurrPONum(Invlinedata) #getFieldFromSource(InvoiceLinedata,'PONumber')
                currStatus = getcurrInvStatus(headerData)
                print("240204: got status --", currStatus) #TEST002 and TEST003 say they are "Approved"????
                if currPONumber[0]!= 'L':
                    NonUMLInvs += 1
                    invoiceType = "NonUML"
                    print("***** NonUML", f)
                    ebCost.removeXMLFile(f, invoiceType)
                else:
                    UMLInvs += 1
                    if currInvNum in ebInvoices:
                        alreadyIn +=1
                        invoiceType = "AlreadyInEB"
                        #print ("Already in EB", currInvNum)
                    else:
                        notAlreadyIn+=1
                        #print ("Not Already in EB", currInvNum)
                        if currPONumber not in activePOs:
                            notActivePO +=1
                            invoiceType = "NonActivePO"
                            print("***** what non active PO???", f)
                            # 240104 This is misleading. It's not a non active PO that is in e-builder. It's a PO that isn't - probably never was
                            # in eb
                            ebCost.removeXMLFile(f, invoiceType)
                        else:
                            activePO +=1
                            currFMP = getFMPNumber(activePOs,currPONumber) #activePOs[str(currPONumber)]['FMP'].decode()
                            print("240204:", currFMP, currStatus)
                            if currStatus == "Paid" or currStatus == "Payable":
                                invoiceType = "EBcost"
                                print("240204: EBcost type, paid or payable")
                                if firstCost:
                                    firstCost = False
                                EBCost +=1
                                #UMLInvs += 1
                                currST = get_Speedtypes(Invlinedata["CustomFieldValueSet"])
                                currFundRule = getFundingRule(currST,fundRules)
                                # We don't want funding rule check!!!
                                if currFundRule == "Funding Rule Not Found":
                                    currFundID = get_XMLfund(Invlinedata["CustomFieldValueSet"])
                                    currFundRule = ebCost.buildFundingRule(currST, currFundID)
                                if currFMP == 'FMP STRING NOT FOUND' and currFundRule == 'Funding Rule Not Found':
                                    nonEB += 1
                                    invoiceType = "nonEB"
                                    #print("****** Got here, removing file", f)
                                    ebCost.removeXMLFile(f, invoiceType)
                                else:
                                    print("******!!!!!!!! looks like EB",currPONumber[0], f)
                                    EB += 1
                                    currVendorInvNum = getCurrVendorInvNumber(headerData)
                                    currLineDesc = getCurrDescription(Invlinedata) #InvoiceLinedata["Item"]["Description"]
                                    currDesc = checkIfDescIsLong(currLineDesc)
                                    currItemAmnt = getCurrItemAmnt(Invlinedata)
                                    currCommItemNum = getCurrCommitItemNum(Invlinedata)
                                    currVoucherID  = getCurrVoucherID(headerData)
                                    currStatus = getcurrInvStatus(headerData)
                                    currApprovedDate = getInvCurrApprovedDate(headerData)
                                    currAmntThisPeriod = getCurrAmntThisPeriod(Invlinedata)
                                    currStoredMaterial = getCurrStoredMaterials()
                                    currQuantityThisPeriod = getCurrQuantityThisPeriod(Invlinedata)

                                    currFMPs.append(currFMP)
                                    currInvNums.append(currInvNum)
                                    currVendorInvoiceNums.append(currVendorInvNum)
                                    currPONums.append(currPONumber)
                                    currDescriptions.append(currDesc)
                                    currInvItemAmnts.append(currItemAmnt)
                                    currCommitItemNums.append(currCommItemNum)
                                    currFundRules.append(currFundRule)
                                    currVoucherIds.append(currVoucherID)
                                    currStatuses.append(currStatus)
                                    currApprovedDates.append(currApprovedDate)
                                    currAmntThisPeriods.append(currAmntThisPeriod)
                                    currStoredMaterials.append(currStoredMaterial)
                                    currQuantityThisPeriods.append(currQuantityThisPeriod)


                                    count+=1

                            elif currStatus == "In Process":
                                invoiceType = "InProcess"
                                InProcess +=1
                            # CHECK THIS: we want Payables to be brought in
                            elif currStatus == "Payable":
                                invoiceType = "Payable"
                                print(currFMP+","+currPONumber+","+currInvNum+","+currStatus+"\n")
                                Payable +=1
                            else:
                                invoiceType = "Other"
                                Other +=1

            else:
                currLineDescList = []
                for k in Invlinedata:

                    currInvNum = getCurrInvNumber(headerData)
                    currPONumber = getCurrPONum(k) #getFieldFromSource(InvoiceLinedata,'PONumber')
                    currStatus = getcurrInvStatus(headerData)
                    if currPONumber[0]!= 'L':
                        NonUMLInvs += 1
                        invoiceType = "NonUML"
                        ebCost.removeXMLFile(f, invoiceType)
                    else:
                        UMLInvs += 1
                        if currInvNum in ebInvoices:
                            alreadyIn +=1
                            invoiceType = "AlreadyInEB"
                            #print ("Already in EB", currInvNum)
                        else:
                            notAlreadyIn+=1
                            #print ("Not Already in EB", currInvNum)
                            if currPONumber not in activePOs:
                                notActivePO +=1
                                invoiceType = "NonActivePO"
                                print("***** what non active PO???", f)
                                # 240104 This is misleading. It's not a non active PO that is in e-builder. It's a PO that isn't - probably never was
                                # in eb
                                ebCost.removeXMLFile(f, invoiceType)
                            else:
                                activePO +=1
                                currFMP = getFMPNumber(activePOs,currPONumber) #activePOs[str(currPONumber)]['FMP'].decode()

                                if currStatus == "Paid" or currStatus == "Payable":
                                    invoiceType = "EBcost"
                                    if firstCost:
                                        firstCost = False
                                    EBCost +=1

                                    currST = get_Speedtypes(k["CustomFieldValueSet"])
                                    # 2 methods for funding rule:
                                    # 1. check existing funding rules in EB
                                    #   - They are not exposed in API unless they've been used?
                                    #   - May not follow naming convention
                                    # 2. Build funding rule from fund description and speedtype
                                    #   - may bring in POs that don't belong (no FMP in PSproject, speedtype like DM
                                    
                                    currFundRule = getFundingRule(currST,fundRules)
                                    if currFundRule == "Funding Rule Not Found":
                                        currFundID = get_XMLfund(Invlinedata["CustomFieldValueSet"])
                                        currFundRule = ebCost.buildFundingRule(currST, currFundID)
                                    if currFMP == 'FMP STRING NOT FOUND' and currFundRule == 'Funding Rule Not Found':
                                        # nonEB += 1
                                        invoiceType = "nonEB"
                                        # ebCost.removeXMLFile(f, invoiceType)
                                    else:
                                        EB += 1
                                        currVendorInvNum = getCurrVendorInvNumber(headerData)
                                        currLineDesc = getCurrDescription(k) #InvoiceLinedata["Item"]["Description"]
                                        currMultiLineDesc = multilineDescription(currLineDesc,currLineDescList)
                                        currDesc = checkIfDescIsLong(currMultiLineDesc)
                                        currItemAmnt = getCurrItemAmnt(k)
                                        currCommItemNum = getCurrCommitItemNum(k)
                                        currVoucherID  = getCurrVoucherID(headerData)
                                        currStatus = getcurrInvStatus(headerData)
                                        currApprovedDate = getInvCurrApprovedDate(headerData)
                                        currAmntThisPeriod = getCurrAmntThisPeriod(k)
                                        currStoredMaterial = getCurrStoredMaterials()
                                        currQuantityThisPeriod = getCurrQuantityThisPeriod(k)

                                        currFMPs.append(currFMP)
                                        currInvNums.append(currInvNum)
                                        currVendorInvoiceNums.append(currVendorInvNum)
                                        currPONums.append(currPONumber)
                                        currDescriptions.append(currDesc)
                                        currInvItemAmnts.append(currItemAmnt)
                                        currCommitItemNums.append(currCommItemNum)
                                        currFundRules.append(currFundRule)
                                        currVoucherIds.append(currVoucherID)
                                        currStatuses.append(currStatus)
                                        currApprovedDates.append(currApprovedDate)
                                        currAmntThisPeriods.append(currAmntThisPeriod)
                                        currStoredMaterials.append(currStoredMaterial)
                                        currQuantityThisPeriods.append(currQuantityThisPeriod)

                                        count+=1

                                elif currStatus == "In Process":
                                    invoiceType = "InProcess"
                                    InProcess +=1
                                elif currStatus == "Payable":
                                    invoiceType = "Payable"
                                    print(currFMP+","+currPONumber+","+currInvNum+","+currStatus+"\n")
                                    Payable +=1
                                else:
                                    invoiceType = "Other"
                                    Other +=1
            print("*****Invoice:", f, "Type:", invoiceType)
            if invoiceType == "EBcost":
                EBCost+=1
            elif invoiceType == "EBProcess":
                EBProcess+=1
            elif invoiceType == "InProcess":
                InProcess+=1
            elif invoiceType == "Payable":
                Payable+=1
            elif invoiceType == "AlreadyInEB":
                AlreadyInEB+=1
            elif invoiceType == "Other":
                Other+=1
            elif invoiceType == "NonUML" or invoiceType == "nonEB":
                ebCost.removeXMLFile(f, invoiceType)

            count += 1

        invoice_data = {1: currFMPs,
                        2: currInvNums,
                        3: currVendorInvoiceNums,
                        4: currPONums,
                        5: currDescriptions,
                        6: currInvItemAmnts,
                        7: currCommitItemNums,
                        8: currFundRules,
                        9: currVoucherIds,
                        10: currStatuses,
                        11: currApprovedDates,
                        12: currAmntThisPeriods,
                        13: currStoredMaterials,
                        14: currQuantityThisPeriods}

    currTime = UMLweb.tstamper2()
    if firstCost == False: # if we found at least one. Can we check count instead?
        
        currWB, currWS = create_Excel()
        write_Headers(currWS, xlHeaders)
        for j in range(1, len(invoice_data) + 1):
            for i in range(len(invoice_data[1])):
                # print(po_data[j][i])
                r = i + 2
                dataincell = str(invoice_data[j][i])
                if len(dataincell) > 365:
                    dataincell = dataincell[0:364]
                    # print (dataincell)
                writeCell(currWS, r, j, dataincell)  # remove this to create seperate files for each XML
                # print(len(dataincell))
        # currWB.save('test1.xlsx')
        # 230627 NEED TO  CHECK IF WE HAVE INVOICES TO OUTPUT
        # look at how we do this for CSV version
        #240204 SAME QUESTION: don't save file if we have nothing to save!
        # what variable tells is how many invoices??????
        oDir = "B:\\dailyImports\\_XML_"
        #oDir = "/Users/kysgattu/FIS/BDrive/dailyImports/_XML_"
        opFile = oDir + currTime + '_InvoicecostImport.xlsx'
        currWB.save(opFile)
        print('Report Saved At: ', oDir + currTime + '_InvoicecostImport.xlsx')
    else:
        print(EBCost)
        opFile = "No output file"
        
    if len(PAYAPFiles) != 0:
        for f in PAYAPFiles:
                theFile = open(f)
                #print(theFile)
                xml_content = theFile.read()
                bwdata = xmltodict.parse(xml_content,encoding='utf-8')
                headerData = bwdata['BuyerInvoiceExportMessage']['Invoice']['InvoiceHeader']
                if 'InvoiceLine' not in bwdata['BuyerInvoiceExportMessage']['Invoice']:
                    print("Invoice Status Cancelled in ", f)
                    #cancelledInvoices +=1

                else:
                    Invlinedata = bwdata['BuyerInvoiceExportMessage']['Invoice']['InvoiceLine']
                    #print(type(InvoiceLinedata))
                    #print(headerData["InvoiceNumber"])
                    if type(Invlinedata) == dict:
                        PAYAPcurrInvNum = getCurrInvNumber(headerData)
                        PAYAPcurrPONumber = getCurrPONum(Invlinedata) #getFieldFromSource(InvoiceLinedata,
                        if PAYAPcurrPONumber[0]!= 'L':
                            ebCost.removeXMLFile(f, lineType="Non UML")
                        else:
                            if PAYAPcurrPONumber in activePOs:
                                PAYAPcurrFMP = getFMPNumber(activePOs,PAYAPcurrPONumber) #activePOs[str(currPONumber)]['FMP'].decode()
                                PAYAPcurrFMPs.append(PAYAPcurrFMP)
                                PAYAPcurrInvNums.append(PAYAPcurrInvNum)
                                PAYAPcurrPONums.append(PAYAPcurrPONumber)
                    else:
                        for k in Invlinedata:
                            PAYAPcurrInvNum = getCurrInvNumber(headerData)
                            PAYAPcurrPONumber = getCurrPONum(k) #getFieldFromSource(InvoiceLinedata,
                            if PAYAPcurrPONumber[0]!= 'L':
                                ebCost.removeXMLFile(f, lineType="Non UML")
                            else:
                                if PAYAPcurrPONumber in activePOs:
                                    PAYAPcurrFMP = getFMPNumber(activePOs,PAYAPcurrPONumber) #activePOs[str(currPONumber)]['FMP'].decode()
                                    PAYAPcurrFMPs.append(PAYAPcurrFMP)
                                    PAYAPcurrInvNums.append(PAYAPcurrInvNum)
                                    PAYAPcurrPONums.append(PAYAPcurrPONumber)

                PAYAP_invoice_data = {1: PAYAPcurrFMPs,
                                      2: PAYAPcurrInvNums,
                                      3: PAYAPcurrPONums}


        currPAYAPWB, currPAYAPWS = create_Excel()
        write_Headers(currPAYAPWS, PAYAPxlHeaders)
        for j in range(1, len(PAYAP_invoice_data) + 1):
            for i in range(len(PAYAP_invoice_data[1])):
                # print(po_data[j][i])
                r = i + 2
                dataincell = str(PAYAP_invoice_data[j][i])
                if len(dataincell) > 365:
                    dataincell = dataincell[0:364]
                    # print (dataincell)
                writeCell(currPAYAPWS, r, j, dataincell)  # remove this to create seperate files for each XML
                # print(len(dataincell))
        # currWB.save('test1.xlsx')
        opPAYAPFile = oDir + currTime + '_PAYAP_InvoicecostImport.xlsx'
        currPAYAPWB.save(opPAYAPFile)
        print('PAYAPReport Saved At: ', oDir + currTime + '_PAYAP_InvoicecostImport.xlsx')

    else:
        opPAYAPFile = "No PAYAP Invoices Found"
    stats = {
        "Timestamp": UMLweb.tstamper(),
        "Source": "XML",
        "EBcost": EBCost,
        "EBprocess": EBProcess,
        "EBCost Type 2.0(PAYAPs)": PAYAPs,
        "InProcess": InProcess,
        "Payable": Payable,
        "Already In EB": AlreadyInEB,
        "Other": Other,
        "Non EB": nonEB
    }

    print('Stats of Invoice Files in the Directory: \n')
    for i in stats:
        print(i + " : " + str(stats[i]))

    # 230701 Changing to appending to an existing HTML or, if none exists, creating a new one
    # 230707 Now, using web_lib, same code we use for CSV/ZIP/Joined processing
    thePath = "B:\\dailyImports\\_InvoiceDataTotals.html"
    COs = {}
    
    UMLweb.outputHTML("Invoice", currTime, stats,COs)

    retVal = {'Invoice_data': stats,
              'stats': stats,
              'invoice_report_excel':opFile,
              'PAYAP_report_excel': opPAYAPFile,
              'stats_html':"SUPERCEDEDED - Check This B:\\dailyImports\\_InvoiceDataTotals.html"}
    return retVal

xlHeaders = {1: "FMP Number",
             2: "Invoice Number",
             3: "Vendor Invoice #",
             4: "Commitment Number",
             5: "Description",
             6: "Invoice Item Amount",
             7: "Commitment Item Number",
             8: "Funding Rule",
             9: "Voucher ID",
             10: "Status",
             11: "Approved Date",
             12: "Amount this Period",
             13: "Stored Materials",
             14: "Quantity This Period" }

PAYAPxlHeaders = {1: "FMP Number",
                  2: "Invoice Number",
                  3: "Commitment Number"}

def main():
    #theDir = "/Users/kysgattu/FIS/BDrive/fromBW/2process/PROCESSED/"
    # theDir = "fromBW16May/"
    theDir = "B:\\fromBW\\2process\\"
    # theDir = "B:\\fromBWLastWeek\\"
    processed_Dir = theDir + "PROCESSED\\"
    InvoiceFiles = glob.glob(theDir + "*_Invoice_*.xml")
    # We always want an html report/ We only want xlsx ONLY if we have at least one EB invoice
    if len(InvoiceFiles) != 0:
        Invoice_dict = checkUMLPAYAP(InvoiceFiles, theDir)
        Invoice_Report = InvXMLtoExcel(Invoice_dict)
    else:
        Invoice_Report = {'Invoice_data': "No Invoice XML Files Found",
                     'stats': "No Invoice XML Files Found",
                     'invoice_report_excel': "No Invoice XML Files Found",
                     'PAYAP_report_excel': "No PAYAP Invoice XML Files Found",
                     'stats_html': "No Invoice XML Files Found"}
        print("Invoice XML Files Not Found!!")

    moveProcessedFiles(theDir, processed_Dir)
    print("Invoice processing complete", Invoice_Report)
    return Invoice_Report

fundRules = eb.get_FundingRules()
ebProjs = eb.get_Projects()
activePOs = eb.get_activePOs(ebProjs)
ebInvoices = eb.get_Invoices()

vendorTypes = {}
build_commitTypes(activePOs)

if __name__ == "__main__":
    main()
