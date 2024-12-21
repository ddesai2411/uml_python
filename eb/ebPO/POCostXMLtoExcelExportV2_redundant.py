#!/usr/bin/env python
# coding: utf-8

import xmltodict, glob
from uml_V2.uml_lib import ebAPI_lib as eb
from uml_V2.uml_lib import ebCostLib as ebCost
from uml_V2.uml_lib import web_lib as UMLweb
# import ebAPI_lib_v2 as eb
from datetime import datetime
from openpyxl import Workbook
import os
import pandas as pd
import shutil



def create_Excel():
    wbOut = Workbook()
    outWS = wbOut.active
    return (wbOut, outWS)

def writeCell(currWS, r, c, val):
    outCell = currWS.cell(row=r, column=c)
    outCell.value = val

def write_Headers(outXL_WS, xlHeaders):
    for h in xlHeaders:
        outCell = outXL_WS.cell(row=1, column=h)
        outCell.value = xlHeaders[h]

def checkUMLPO(POFiles, filepath):
    PO_dict = {'UMLPOFiles': [], 'nonUMLPOFiles': []}
    # UMLPOFiles = []
    # nonUMLPOFiles = []
    for i in range(len(POFiles)):
        if POFiles[i].startswith(filepath + 'Jaggaer_PO_L'):
            PO_dict['UMLPOFiles'].append(POFiles[i])
            # UMLPOFiles.append(POFiles[i])
        else:
            PO_dict['nonUMLPOFiles'].append(POFiles[i])
    return PO_dict

def deleteNonUML(PO_dict):
    nonUML = PO_dict['nonUMLPOFiles']
    for f in nonUML:
        #print(f)
        if os.path.exists(f):
            # Delete the file
            os.remove(f)
            #print("File deleted successfully!")
        else:
            ignoreThis = 1
            #print("The file does not exist.")

def moveProcessedFiles(source_Dir, processed_Dir):
    POFiles = glob.glob(source_Dir + "*_PO_*.xml")
    for i in range(len(POFiles)):
        POFile = POFiles[i]
        POFileName = POFiles[i].split('\\')[-1]
        #print(POFile)
        move_dest = os.path.join(processed_Dir, POFileName)
        # print(move_dest)
        if not os.path.exists(move_dest):
            #print(">>>> Move "< processed_Dir, POFileName)
            #print('e')
            shutil.move(POFile, processed_Dir)
        #else:
            #print("Not moving, already tehre??")

def get_FMPfromPSproject(psProj):
    retVal = "FMP NOT FOUND"
    #print (">>>>", psProj)
    toks = psProj.split("FMP")
    # print(toks)
    if len(toks) < 2:
        #print("NO FMP")
        retVal = "NO FMP -- O&S? DO NOT BUILD FUNDING RULE"
    else:
        #print(toks[1])
        if toks[1][0]=="0":
            retVal = toks[1][1:]
        else:
            retVal = toks[1]
    return retVal

def get_XMLproject(theData):
    retVal = "PS Project Not found"
    for f in theData:
        if f["@name"] == "Project":
            if type(f["CustomFieldValue"]) == dict:
                currPSproj = f["CustomFieldValue"]["Value"][:-2]
                #print(currPSproj)
                return currPSproj
            else:
                multiProj = []
                for i in range(len(f["CustomFieldValue"])):
                    #print(f["CustomFieldValue"][i]["Value"])
                    currPSproj = f["CustomFieldValue"][i]["Value"][:-2]
                    if currPSproj not in multiProj:
                        multiProj.append(currPSproj)
                    retVal = multiProj
                PSprojects = ",".join(multiProj)
                return PSprojects
    return retVal

def get_XMLfund(theData):
    for f in theData:
        if f["@name"] == "Fund":
            #print("GOT IT")
            if type(f["CustomFieldValue"]) == dict:
                currFund = f["CustomFieldValue"]["Value"][:-2]
                #print(currFund)
                return currFund
            else:
                multiFund = []
                for i in range(len(f["CustomFieldValue"])):
                    #print(f["CustomFieldValue"][i]["Value"])
                    currFund = f["CustomFieldValue"][i]["Value"][:-2]
                    if currFund not in multiFund:
                        multiFund.append(currFund)
                    retVal = multiFund
                PSprojects = ",".join(multiFund)
                return PSprojects

def get_Speedtypes(theData):
    for f in theData:
        if f["@name"] == "Speedtype":
            if type(f["CustomFieldValue"]) == dict:
                currST = f["CustomFieldValue"]["Value"][:-2]
                #print(currST)
                return currST
            else:
                multiST = []
                for i in range(len(f["CustomFieldValue"])):
                    #print(f["CustomFieldValue"][i]["Value"])
                    currST = f["CustomFieldValue"][i]["Value"][:-2]
                    if currST not in multiST:
                        multiST.append(currST)
                    retVal = multiST
                speedtypes = ",".join(multiST)
                #for i in range(len(multiST)):
                    #speedtypes = speedtypes + "," + multiST[i]
                return speedtypes

def get_customField(theData, theField):
    retVal = "NOT FOUND"

    for f in theData:
        # print "???", f["@name"]
        if f["@name"] == theField:
            try:
                retVal = f["CustomFieldValue"]["Value"]
            except:
                retVal = "MULTIPLE"
                # for g in f:
                #    print( "\tget_customField", g)
    # print(retVal)
    return retVal

def getFieldFromSource(source, field):
    retval = ''
    if source and source != 'Field Not Found in XML' and field in source:
        # retval = retval + source[field]
        fieldValue = source[field]
        if fieldValue != None:
            retval = fieldValue
            return retval
        else:
            retval = retval + 'Field Not Found in XML'
            return retval
    else:
        retval = retval + 'Field Not Found in XML'
        return retval

# --------------------------------------------------------
# These Methods are not used but kept as a backup - Getting FMPs using SpeedTypes instead
def checkFMP(theFMP):
    retVal = True
    for c in theFMP:
        if c not in "0123456789":
            retVal = False
            break
    return retVal

# --------------------------------------------------------
def getFMPNumber(activePOs,currPONumber):
    currFMP = activePOs[str(currPONumber)]['FMP'].decode()
    return currFMP

def dateStrToDate(currCommitDate):
    retval = ''
    yrStr = currCommitDate[0:4]
    moStr = currCommitDate[5:7]
    dtStr = currCommitDate[8:10]
    retval = moStr + '/' + dtStr + '/' + yrStr
    return retval

def getCurrFMPLead(currFMP,ebProjs):
    if "," in currFMP:
        currLead = "TBD"
    else:
        currLead = ebProjs[currFMP]["FMlead"] #ebProjs return FMP values as Bytes, hence encoding.
    return currLead

def build_commitTypes(activePOs):
    global vendorTypes

    for p in activePOs:
        currVendor = activePOs[p]["Vendor"]  # .encode("utf-8")
        currType = activePOs[p]["CommitmentType"]  # .encode("utf-8")

        if currVendor not in vendorTypes:
            vendorTypes[currVendor] = [currType]
        else:
            if currType not in vendorTypes[currVendor]:
                vendorTypes[currVendor].append(currType)
    vendorTypes["DAEDALUS PROJECTS INC"] = ["4. Purchase Order-House Doctors & OPM's"]
    vendorTypes["DIAMOND RELOCATION INC"] = ["3. Purchase Order-Master Service Vendor/Contractor"]
    vendorTypes["EXPRESS SIGN & GRAPHICS INC"] = ["3. Purchase Order-Master Service Vendor/Contractor"]
    vendorTypes["MILLER DYER SPEARS INC"] = ["4. Purchase Order-House Doctors & OPM's"]
    vendorTypes["MODULEX NEW ENGLAND"] = ["3. Purchase Order-Master Service Vendor/Contractor"]
    vendorTypes["R-SQUARED OFFICE PANELS & FURNITURE INC"] = ["3. Purchase Order-Master Service Vendor/Contractor"]
    vendorTypes["STATE ELECTRIC CORPORATION"] = ["3. Purchase Order-Master Service Vendor/Contractor"]
    vendorTypes["TRC ENVIRONMENTAL CORPORATION"] = ["1. Purchase Order"]
    vendorTypes["TRIUMVIRATE ENVIRONMENTAL INC"] = ["1. Purchase Order"]
    vendorTypes["VAV INTL INC"] = ["4. Purchase Order-House Doctors & OPM's"]
    vendorTypes["WILLIAM LOWE & SONS CORP"] = ["3. Purchase Order-Master Service Vendor/Contractor"]

# 1. Commitment Number
def getCurrPONum(datasource):
    currPONum = getFieldFromSource(datasource, 'PONumber')
    return currPONum

# 2. FMP Number
def getFMPFromST(currST, fundingRule):
    retval = ''
    flag = currST.find(',')
    if flag == -1:
        if currST in fundingRule.keys():
            FMPs = fundingRule[currST]['FMP']
            for i in range(len(FMPs)):
                currFMPs = ",".join(FMPs)
            retval = retval + currFMPs
        else:
            retval = retval + "FMP STRING NOT FOUND"
        return retval
    else:
        FMPlist = []
        speedtypes = currST.split(',')
        for i in speedtypes:
            #print(i)
            if i in fundingRule.keys():
                #fundrules = ','.join(fundingRule[i]['Name'])
                FMPlist.append(fundingRule[i]['FMP'])
        FMPs = []
        if len(FMPlist) != 0:
            for i in range(len(FMPlist)):
                for j in range(len(FMPlist[i])):
                    retval += FMPlist[i][j] + ','
            retval = retval[:-1]
        else:
            retval = "FMP STRING NOT FOUND"
        return retval

def getFMPFromExternalNote(dataSource):
    externalInfo = getFieldFromSource(dataSource,'ExternalInfo')
    # print("::",externalInfo)
    externalNote = getFieldFromSource(externalInfo,'Note')
    retVal = ebCost.parseHeaderNotes(externalNote)
    return retVal

def getFMPFromDescripion(dataSource):
    item = getFieldFromSource(dataSource, 'Item')
    description = getFieldFromSource(item, 'Description')
    retVal = ebCost.parse_Buyways_Description(description)
    return retVal


# 3. Commitment Type
def getCommitTypeFromVendor(vendorName):
    currVendor = vendorName.upper()
    # global vendorTypes
    retVal = "NOT FOUND"
    # print type(currVendor)
    if currVendor in vendorTypes:
        # print "Found!", currVendor
        # print "+++++++++++++++++++++++++++++++++", vendorTypes[currVendor]
        try:
            if len(vendorTypes[currVendor]) == 1:
                retVal = vendorTypes[currVendor][0]
            else:
                retVal = "CHECK: MULTIPLE VALUES"
        except:
            ignoreThis = 1
            # print ".......failed on vendorTypes"
    """
    try:
        retVal = vendorTypes[v]
        if len(vendorTypes[v]) > 1:
            retVal = "CHECK TYPE: MULTIPLE VALUES"
    except:
        print ">?>?", currVendor
        retVal = "VENDOR NOT FOUND"
    """
    # print "from getCommitTypeFromVendor:\n", retVal
    return retVal

def getCurrCommitType(currPONumber,currVendor):
    if currPONumber in activePOs:
        retval = activePOs[currPONumber]['CommitmentType']
    else:
        retval = getCommitTypeFromVendor(currVendor)
    return retval
# Unused
def getCommitType(currCompName,FMlead):
    vendorName = currCompName.upper()
    if FMlead == "O&S":
        return "O&S: DO NOT IMPORT"
    else:
        return getCommitTypeFromVendor(vendorName)


# 4. Company Name
def getCurrCompName(datasource):
    supplier = getFieldFromSource(datasource, 'Supplier')
    supplier_name = getFieldFromSource(supplier, 'Name')
    return supplier_name

# 5. Budget Line Item
def getBudgetLineItem(datasource):
    budgetLineItem = get_customField(datasource["CustomFieldValueSet"], "Account")
    budgetTasks = {"772130": "20.772130",
                   "772140": "40.772140",
                   "772160": "30.772160",
                   "772180": "60.772180",
                   "772220": "30.772220",
                   "772230": "30.772230",
                   "734200": "60.734200",
                   "734310": "10.734310",
                   "734800": "50.734800",
                   "735100": "10.735100",
                   "735200": "80.735200",
                   "735600": "10.735600",
                   "740800": "60.740800",
                   "741970": "60.741970",
                   "741980": "50.741980",
                   "744300": "60.744300",
                   "744700": "30.744700",
                   "757060": "20.757060",
                   "757120": "20.757120",
                   "757190": "10.757190",
                   "757220": "70.757220",
                   "761080": "50.761080",
                   "761090": "60.761090",
                   "761210": "40.761210",
                   "761211": "40.761211",
                   "761460": "10.761460",
                   "761590": "70.761590",
                   "761600": "30.761600",
                   "761603": "30.761603",
                   "761604": "30.761604",
                   "763100": "50.763100",
                   "763120": "50.763120",
                   "763140": "50.763140",
                   "763210": "60.763210",
                   "763400": "30.763400",
                   "763900": "60.763900",
                   "764300": "50.764300",
                   "766100": "50.766100",
                   "766300": "60.766300",
                   "766600": "60.766600",
                   "767210": "60.767210",
                   "767300": "60.767300",
                   "767800": "60.767800",
                   "768100": "50.768100",
                   "768210": "60.768210",
                   "768300": "60.768300",
                   "768600": "60.768600",
                   "772010": "20.772010",
                   "772050": "10.772050",
                   "772060": "20.772060",
                   "772120": "10.772120",
                   "772130": "20.772130",
                   "772140": "40.772140",
                   "772160": "30.772160",
                   "772180": "60.772180",
                   "772220": "30.772220",
                   "772230": "30.772230",
                   "700T10": "10.700T10",
                   "700T20": "20.700T20",
                   "700T30": "30.700T30",
                   "700T40": "40.700T40",
                   "700T50": "50.700T50",
                   "700T60": "60.700T60",
                   "700T70": "70.700T70",
                   "700T80": "80.700T80",
                   "99CONT": "99.99CONT"}
    if budgetLineItem[:-2] in budgetTasks:
        currBudgetLineItem = budgetTasks[budgetLineItem[:-2]]
    else:
        currBudgetLineItem = budgetTasks['99CONT']
    return currBudgetLineItem

# 6. Status - Always Approved (draft was more dangerous!)

# 7. Item Number
def getPOCurrItemNum(datasource):
    linenumber = getFieldFromSource(datasource, '@linenumber')
    return linenumber

# 8. Description
def getCurrDescription(datasource):
    item = getFieldFromSource(datasource, 'Item')
    description = getFieldFromSource(item, 'Description')
    return description

def multilineDescription(currDesc, currLineDescList):
    # curr_list = ['']
    curr_desc = ''
    if currDesc in currLineDescList:
        # print('check')
        curr_desc = ''.join(currDesc)
        currLineDescList.append(currDesc)
        return curr_desc
    else:
        # print('here')
        currLineDescList.append(currDesc)
        # print(currLineDescList)
        curr_desc = ' | '.join(str(item) for item in currLineDescList)
        return curr_desc

def checkIfDescIsLong(currDescription):
    if len(currDescription) > 365:
        retval = "SEE BUYWAYS FOR FULL DESCRIPTIONS - " + currDescription
    else:
        retval = currDescription
    return retval

# 9. Item Quantity
def getCurrItemQuantity(datasource):
    quantity = getFieldFromSource(datasource, 'Quantity')
    return quantity

# 10. Item Unit Cost
def getPOCurrItemCost(datasource):
    linecharge = getFieldFromSource(datasource, 'LineCharges')
    unitprice = getFieldFromSource(linecharge, 'UnitPrice')
    money = getFieldFromSource(unitprice, 'Money')
    itemcost = getFieldFromSource(money, '#text')
    return itemcost

# 11. Commitment Date
def getPOCurrCommitDate(datasource):
    dateStr = getFieldFromSource(datasource, 'CreateDateTime')
    date = dateStrToDate(dateStr)
    return date

# 12. Company Number
def getCurrCompNum(datasource):
    supplier = getFieldFromSource(datasource, 'Supplier')
    supplier_num = getFieldFromSource(supplier, 'SupplierNumber')
    return supplier_num

# 13. Commitment Item Amount
def getPOCurrCommitItemAmnt(datasource):
    linecharge = getFieldFromSource(datasource, 'LineCharges')
    extendedprice = getFieldFromSource(linecharge, 'ExtendedPrice')
    money = getFieldFromSource(extendedprice, 'Money')
    commitItemAmount = getFieldFromSource(money, '#text')
    return commitItemAmount

# 14. Item Unit Of Measure
def getcurrUnitOfMeasure(datasource):
    item = getFieldFromSource(datasource, 'Item')
    prodUnitOfMeasure = getFieldFromSource(item, 'ProductUnitOfMeasure')
    measurement = getFieldFromSource(prodUnitOfMeasure[1], 'Measurement')
    measurementUnit = getFieldFromSource(measurement, 'MeasurementUnit')
    return measurementUnit


# Not used but kept as backup functionality
def getCurrPOItemUnitMeasure(datasource, currItemCost):
    item = getFieldFromSource(datasource, 'Item')
    prodUnitOfMeasure = getFieldFromSource(item, 'ProductUnitOfMeasure')
    measurement = getFieldFromSource(prodUnitOfMeasure[1], 'Measurement')
    measurementUnit = getFieldFromSource(measurement, 'MeasurementUnit')
    itemUnitMeasure = getPOItemUnitMeasure(measurementUnit, currItemCost)
    return itemUnitMeasure

def getPOItemUnitMeasure(itemUnit, currItemCost):
    # value = getFieldFromSource(POlinedata["Item"]['ProductUnitOfMeasure'][1]['Measurement'],'MeasurementUnit')
    if itemUnit != "Field Not Found in XML":
        if itemUnit.find(" ") == -1:
            measureValue = currItemCost + '/' + itemUnit
            return measureValue
        else:
            split_value = itemUnit.split(" ")
            if split_value[0] == '0.0' or split_value[0] == '0.00' or split_value[0] == '0':
                # print('checl')
                split_value[0] = currItemCost
            # else:
            # print(1.1)
            # print(split_value)
            measurevalue = '/'.join(split_value)
            # print(split_value)
            return measurevalue
    else:
        return itemUnit

# 15. Funding Rules
def getFundingRule(currST, fundingRule):
    #print(currST)
    retval = ''
    flag = currST.find(',')
    if flag == -1:
        if currST in fundingRule.keys():
            retval = retval + fundingRule[currST]['Name']
        else:
            retval = retval + "Funding Rule Not Found"
        return retval
    else:
        fundrules = []
        speedtypes = currST.split(',')
        #print(speedtypes)
        for i in speedtypes:
            #print(i)
            if i in fundingRule.keys():
                #fundrules = ','.join(fundingRule[i]['Name'])
                fundrules.append(fundingRule[i]['Name'])
        #print(fundrules)
        if len(fundrules) != 0:
            for i in range(len(fundrules)):
                retval = ",".join(fundrules)
        else:
            retval = "Funding Rule Not Found"
        #print(retval)
        return retval

# 16. Speedtype
def get_Speedtypes(theData):
    for f in theData:
        if f["@name"] == "Speedtype":
            if type(f["CustomFieldValue"]) == dict:
                currST = f["CustomFieldValue"]["Value"][:-2]
                #print(currST)
                return currST
            else:
                multiST = []
                for i in range(len(f["CustomFieldValue"])):
                    #print(f["CustomFieldValue"][i]["Value"])
                    currST = f["CustomFieldValue"][i]["Value"][:-2]
                    if currST not in multiST:
                        multiST.append(currST)
                    retVal = multiST
                speedtypes = ",".join(multiST)
                #for i in range(len(multiST)):
                #speedtypes = speedtypes + "," + multiST[i]
                return speedtypes

# 17. Peoplesoft PO# - Same as PO Number

# 18. Origin Code
def getCurrOriginCode(datasource):
    currOriginCode = get_customField(datasource["CustomFieldValueSet"], "Origin Code")
    return currOriginCode

# 19. Commodity Code
def getCurrCommCode(datasource):
    item = getFieldFromSource(datasource, 'Item')
    commodityCode = getFieldFromSource(item, 'CommodityCode')
    return commodityCode

# 20. Retainage Amount - If Commitment Type is type 2 - Retainage amount will be 5.0
def getCurrRetainageAmnt(currCommType):
    if currCommType[0] == '2':
        return '5.0'
    else:
        return '0.0'


# 21. Notes
def getcurrNotes(currST):
    flag = currST.find(',')
    if flag == -1:
        retval = ''
    else:
        retval = "Multiple SpeedTypes" + str(currST)
    return retval


def POXMLtoExcel(PO_dict):
    currPONums = []
    currFMPs = []
    currCommTypes = []
    currCompNames = []
    currBudgLineItems = []
    currStatuses = []
    currItemNums = []
    currDescriptions = []
    currItemQuantities = []
    currItemCosts = []
    currCommitDates = []
    currCompNums = []
    currCommitItemAmnts = []
    currItemUnitMeasures = []
    currFundRules = []
    currSTs = []
    currPSPOs = []
    currOriginCodes = []
    currCommCodes = []
    currRetainageAmnts = []
    currNotes = []
    count = 0
    nonEB = 0
    EBcost = 0
    EBProcess = 0
    EBexists = 0
    EBexistsCO = 0
    lineType = ''

    EB = 0
    alreadyIn = 0 # This also satisfies the condition ebExists in Old Code
    ff =0

    ebST = eb.get_FundingRules()

    UMLPOFiles = PO_dict['UMLPOFiles']
    NonUMLPOFiles = PO_dict['nonUMLPOFiles']
    UMLPOs = len(UMLPOFiles)
    NonUMLPOs = len(NonUMLPOFiles)
    COs = {'PO Number':'FMP Number'}

    counts = {'Timestamp': '',
             'Source': "XML",
             'EBprocess': 0,
             'EBcost': 0,
             'EBexists': 0,
             'EBexistsCO?': 0,
             'nonEB': 0,
             'nonUML': NonUMLPOs,
             }
    changeOrders = {"PONums":"FMPs"}

    for f in UMLPOFiles:
        # print(f)
        theFile = open(f,encoding='utf-8')
        count += 1
        # print(count,f)
        xml_content = theFile.read()
        bwdata = xmltodict.parse(xml_content, encoding='utf-8')
        headerData = bwdata['PurchaseOrderMessage']['PurchaseOrder']['POHeader']
        POlinedata = bwdata['PurchaseOrderMessage']['PurchaseOrder']['POLine']

        # check if the PO belongs to E-Builder
        if type(POlinedata) == dict:
            #getFieldFromSource(source, field)

            currST = get_Speedtypes(POlinedata["CustomFieldValueSet"])
            currProj = get_XMLproject(POlinedata["CustomFieldValueSet"])
            currPSFMP = get_FMPfromPSproject(currProj) # does not handle multiple project numbers??
            currDescFMP = getFMPFromDescripion(POlinedata)
            currHNFMP = getFMPFromExternalNote(headerData)
            currFundRule = getFundingRule(currST, fundRules)
            currPONumber = currPSPO = getCurrPONum(headerData)
            currOriginCode = getCurrOriginCode(POlinedata)
            currCompName = getCurrCompName(headerData)  # headerData['Supplier']['Name']
            currCommType = getCurrCommitType(currPONumber, currCompName)
            currBudgLineItem = getBudgetLineItem(POlinedata)
            currStatus = "Approved"
            currItemNum = getPOCurrItemNum(POlinedata)  # POlinedata["@linenumber"]
            currLineDesc = getCurrDescription(POlinedata)  # POlinedata["Item"]["Description"]
            currDesc = checkIfDescIsLong(currLineDesc)
            currItemQuan = getCurrItemQuantity(POlinedata)  # POlinedata["Quantity"]
            currItemCost = getPOCurrItemCost(POlinedata)  # POlinedata["LineCharges"]["UnitPrice"]["Money"]["#text"]
            currCommitDate = getPOCurrCommitDate(headerData)  # dateStrToDate(headerData['CreateDateTime'])
            currCompNum = getCurrCompNum(headerData)  # headerData['Supplier']['@id']
            currCommitItemAmnt = getPOCurrCommitItemAmnt(POlinedata)  # POlinedata["LineCharges"]["ExtendedPrice"]["Money"]["#text"]
            # ItemUnitMeasure = POlinedata["Item"]['ProductUnitOfMeasure'][1]['Measurement']['MeasurementUnit']
            # itemUnit = getFieldFromSource(POlinedata["Item"]['ProductUnitOfMeasure'][1]['Measurement'],'MeasurementUnit')
            currItemUnitMeasure = getcurrUnitOfMeasure(POlinedata)  # getCurrPOItemUnitMeasure(POlinedata, currItemCost)
            currCommCode = getCurrCommCode(POlinedata)  # POlinedata["Item"]["CommodityCode"]
            currRetainageAmnt = getCurrRetainageAmnt(currCommType)
            curr_Notes = getcurrNotes(currST)

            if currOriginCode == "LEB":
                lineType = "EBprocess"
            else:
                if currPSFMP != "NO FMP -- O&S? DO NOT BUILD FUNDING RULE" and currPSFMP != "FMP NOT FOUND":
                    lineType = "EBcostFMP"
                    print('here')
                    currFMP = currPSFMP
                elif currHNFMP != "NON FMP":
                    lineType = "EBcostFMP"
                    currFMP = currHNFMP
                elif currDescFMP != "NON FMP":
                    lineType = "EBcostFMP"
                    currFMP = currDescFMP
                else:
                    lineType = "nonEB"
                    currFMP = "NON FMP"
                    # Checking whether the Speedtype belongs to E-Builder
                    if currST in ebST:
                        print(f"For the PO Number {currPONumber}, Speed Type {currST} belongs to E-Builder")
                    else:
                        print(f"For the PO Number {currPONumber}, Speed Type {currST} does not belong to E-Builder")
                # print(currFMP)
            if currFundRule == "Funding Rule Not Found" and len(currFMP) > 6:
                currFundID = get_XMLfund(POlinedata["CustomFieldValueSet"])
                currFundRule = ebCost.buildFundingRule(currST, currFundID)
            # print(lineType)
            if lineType != "nonEB":
                if currPONumber in activePOs:
                    lineType = "EBexists"
                    currBWValue = float(getPOCurrCommitItemAmnt(POlinedata))  # POlinedata["LineCharges"]["ExtendedPrice"]["Money"]["#text"]
                    currEBValue = float(activePOs[currPONumber]["Value"])
                    delta = currEBValue - currBWValue
                    # if delta < 0  or delta > 0:
                    if currEBValue != currBWValue:
                        # print('Equal')
                        lineType = "EBexistsCO?"
                        changeOrders[currPONumber] = currFMP

                # print(lineType)
                if lineType == "EBcostFMP":
                    currPONums.append(currPONumber)
                    currFMPs.append(currFMP)
                    currCommTypes.append(currCommType)
                    currCompNames.append(currCompName)
                    currBudgLineItems.append(currBudgLineItem)
                    currStatuses.append(currStatus)
                    currItemNums.append(currItemNum)
                    currDescriptions.append(currDesc)
                    currItemQuantities.append(currItemQuan)
                    # currItemCosts.append(currItemCost)
                    currItemCosts.append("{:,.2f}".format(float(currItemCost)))
                    currCommitDates.append(currCommitDate)
                    currCompNums.append(currCompNum)
                    currCommitItemAmnts.append("{:,.2f}".format(float(currCommitItemAmnt)))
                    currItemUnitMeasures.append(currItemUnitMeasure)
                    currFundRules.append(currFundRule)
                    currSTs.append(currST)
                    currPSPOs.append(currPSPO)
                    currOriginCodes.append(currOriginCode)
                    currCommCodes.append(currCommCode)
                    currRetainageAmnts.append(currRetainageAmnt)
                    currNotes.append(curr_Notes)

        else:
            currLineDescList = []
            for k in POlinedata:
                currST = get_Speedtypes(k["CustomFieldValueSet"])
                currProj = get_XMLproject(k["CustomFieldValueSet"])
                currPSFMP = get_FMPfromPSproject(currProj)  # does not handle multiple project numbers??
                currDescFMP = getFMPFromDescripion(k)
                currHNFMP = getFMPFromExternalNote(headerData)
                currFundRule = getFundingRule(currST, fundRules)
                currPONumber = currPSPO = getCurrPONum(headerData)
                currOriginCode = getCurrOriginCode(k)
                currCompName = getCurrCompName(headerData)  # headerData['Supplier']['Name']
                currCommType = getCurrCommitType(currPONumber, currCompName)
                currBudgLineItem = getBudgetLineItem(k)
                currStatus = "Approved"
                currItemNum = getPOCurrItemNum(k)  # k["@linenumber"]
                currLineDesc = getCurrDescription(k)  # k["Item"]["Description"]
                currDesc = checkIfDescIsLong(currLineDesc)
                currItemQuan = getCurrItemQuantity(k)  # k["Quantity"]
                currItemCost = getPOCurrItemCost(k)  # k["LineCharges"]["UnitPrice"]["Money"]["#text"]
                currCommitDate = getPOCurrCommitDate(headerData)  # dateStrToDate(headerData['CreateDateTime'])
                currCompNum = getCurrCompNum(headerData)  # headerData['Supplier']['@id']
                currCommitItemAmnt = getPOCurrCommitItemAmnt(
                    k)  # k["LineCharges"]["ExtendedPrice"]["Money"]["#text"]
                # ItemUnitMeasure = k["Item"]['ProductUnitOfMeasure'][1]['Measurement']['MeasurementUnit']
                # itemUnit = getFieldFromSource(k["Item"]['ProductUnitOfMeasure'][1]['Measurement'],'MeasurementUnit')
                currItemUnitMeasure = getcurrUnitOfMeasure(
                    k)  # getCurrPOItemUnitMeasure(k, currItemCost)
                currCommCode = getCurrCommCode(k)  # k["Item"]["CommodityCode"]
                currRetainageAmnt = getCurrRetainageAmnt(currCommType)
                curr_Notes = getcurrNotes(currST)

                if currOriginCode == "LEB":
                    lineType = "EBprocess"
                else:
                    if currPSFMP != "NO FMP -- O&S? DO NOT BUILD FUNDING RULE" and currPSFMP != "FMP NOT FOUND":
                        lineType = "EBcostFMP"
                        currFMP = currPSFMP
                    elif currHNFMP != "NON FMP":
                        lineType = "EBcostFMP"
                        currFMP = currHNFMP
                    elif currDescFMP != "NON FMP":
                        lineType = "EBcostFMP"
                        currFMP = currDescFMP
                    else:
                        lineType = "nonEB"
                        currFMP = "NON FMP"
                        # Checking whether the Speedtype belongs to E-Builder
                        if currST in ebST:
                            print(f"For the PO Number {currPONumber}, Speed Type {currST} belongs to E-Builder")
                        else:
                            print(f"For the PO Number {currPONumber}, Speed Type {currST} does not belong to E-Builder")

                if currFundRule == "Funding Rule Not Found" and len(currFMP) > 6:
                    currFundID = get_XMLfund(k["CustomFieldValueSet"])
                    currFundRule = ebCost.buildFundingRule(currST, currFundID)
                # print(lineType)
                if lineType != "nonEB":
                    if currPONumber in activePOs:
                        lineType = "EBexists"
                        currBWValue = float(getPOCurrCommitItemAmnt(
                            k))  # k["LineCharges"]["ExtendedPrice"]["Money"]["#text"]
                        currEBValue = float(activePOs[currPONumber]["Value"])
                        delta = currEBValue - currBWValue
                        if delta < 0 or delta > 0:
                            # print('Equal')
                            lineType = "EBexistsCO?"
                            changeOrders[currPONumber] = currFMP

                    # print(lineType)
                    if lineType == "EBcostFMP":
                        currPONums.append(currPONumber)
                        currFMPs.append(currFMP)
                        currCommTypes.append(currCommType)
                        currCompNames.append(currCompName)
                        currBudgLineItems.append(currBudgLineItem)
                        currStatuses.append(currStatus)
                        currItemNums.append(currItemNum)
                        currDescriptions.append(currDesc)
                        currItemQuantities.append(currItemQuan)
                        # currItemCosts.append(currItemCost)
                        currItemCosts.append("{:,.2f}".format(float(currItemCost)))
                        currCommitDates.append(currCommitDate)
                        currCompNums.append(currCompNum)
                        currCommitItemAmnts.append("{:,.2f}".format(float(currCommitItemAmnt)))
                        currItemUnitMeasures.append(currItemUnitMeasure)
                        currFundRules.append(currFundRule)
                        currSTs.append(currST)
                        currPSPOs.append(currPSPO)
                        currOriginCodes.append(currOriginCode)
                        currCommCodes.append(currCommCode)
                        currRetainageAmnts.append(currRetainageAmnt)
                        currNotes.append(curr_Notes)

        # print(lineType)
        # if lineType == "EBcostFMP" or "EBcostST":
        #     counts["EBcost"] += 1
        # else:
        #     counts[lineType] += 1

        if lineType == "EBcostFMP" or lineType == "EBcostST":
            # print("CCCC")
            counts["EBcost"] += 1
        elif lineType == "EBProcess":
            # print("PPPP")
            counts["EBprocess"] += 1
        elif lineType == "EBexists":
            # print("EEEE")
            counts['EBexists'] += 1
        elif lineType == "EBexistsCO?":
            # print("ECEC")
            if currPONumber not in COs:
                COs[currPONumber] = currFMP
            counts['EBexistsCO?'] += 1
        elif lineType == "nonEB":
            # print("nonEB")
            counts['nonEB'] += 1



        po_data = {1: currPONums,
                       2: currFMPs,
                       3: currCommTypes,
                       4: currCompNames,
                       5: currBudgLineItems,
                       6: currStatuses,
                       7: currItemNums,
                       8: currDescriptions,
                       9: currItemQuantities,
                       10: currItemCosts,
                       11: currCommitDates,
                       12: currCompNums,
                       13: currCommitItemAmnts,
                       14: currItemUnitMeasures,
                       15: currFundRules,
                       16: currSTs,
                       17: currPSPOs,
                       18: currOriginCodes,
                       19: currCommCodes,
                       20: currRetainageAmnts,
                       21: currNotes}


    currWB, currWS = create_Excel()
    currTime = UMLweb.tstamper2()
    write_Headers(currWS, xlHeaders)
    for j in range(1, len(po_data) + 1):
        for i in range(len(po_data[1])):
            # print(po_data[j][i])
            r = i + 2
            dataincell = str(po_data[j][i])
            if len(dataincell) > 365:
                dataincell = dataincell[0:364]
                # print (dataincell)
            writeCell(currWS, r, j, dataincell)  # remove this to create seperate files for each XML
            # print(len(dataincell))
    # currWB.save('test1.xlsx')

    # oDir = "B:\\dailyImports\\_XML_"
    oDir = "/Users/kysgattu/FIS/BDrive/dailyImports/_XML_/"
    # oDir = "C:\\Users\\K_Gattu\\PycharmProjects\\uml_V2\\umlV2\\outputfiles\\"
    opFile = oDir + currTime + '_POcostImport.xlsx'
    currWB.save(opFile)
    print('Report Saved At: ', oDir + currTime + '_POcostImport.xlsx')

    print("Counts of Each Type of POs:\n",counts)
    for i in counts:
        print(i + ":" + str(counts[i]))

    retVal = {"PO Data":po_data,
              'Stats':counts,
              'ChangeOrders': changeOrders,
              'ImportFile': opFile}

    return retVal

xlHeaders = {1: "Commitment Number",
             2: "FMP Number",
             3: "Commitment Type",
             4: "Company Name",
             5: "Budget Line Item",
             6: "Status",
             7: "Item Number",
             8: "Description",
             9: "Item Quantity",
             10: "Item Unit Cost",
             11: "Commitment Date",
             12: "Company Number",
             13: "Commitment Item Amount",
             14: "Item Unit of Measure",
             15: "Funding Rule",
             16: "Speedtype",
             17: "PeopleSoft PO#",
             18: "Origin Code",
             19: "Commodity Code",
             20: "Retainage Amount",
             21: "Notes"
             }

def main():
    # theDir = "DataFiles/testdata/"
    # theDir = "CostXML_IP_OP/fromBW/"
    # theDir = "B:\\fromBW\\"
    #theDir = "C:\\temp\\"
    # theDir = "B:\\fromBWLastWeek\\"
    theDir = "/Users/kysgattu/FIS/BDrive/fromBW/"
    processed_Dir = theDir + "PROCESSED/"
    POFiles = glob.glob(theDir + "*_PO_*.xml")
    if len(POFiles) != 0:
        PO_dict = checkUMLPO(POFiles, theDir)
        PO_Report = POXMLtoExcel(PO_dict)
        deleteNonUML(PO_dict)
    else:
        PO_Report = {'PO Data': "No PO XML Files Found!!",
                     'Stats': "No PO XML Files Found!!",
                     'ChangeOrders': "No PO XML Files Found!!",
                     'ImportFile': "No PO XML Files Found!!"}
        print("No PO XML Files Found!!")
    # moveProcessedFiles(theDir,processed_Dir)
    return PO_Report

fundRules = eb.get_FundingRules()
ebProjs = eb.get_Projects()
activePOs = eb.get_activePOs(ebProjs)

vendorTypes = {}
build_commitTypes(activePOs)

if __name__ == "__main__":
    main()





































