from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os
import pandas as pd

# checkBWcsv
# PURPOSE:  Determine if a CSV file appears to be a valid Buyways PO or Voucher **transaction** report.
# METHOD:   Compare 1st line of CSV with standard Buyways PO Header and Voucher header. Rewind CSV to beginning for parsing
# INPUT:    file path
# OUTPUT:   Category/triage - Is a PO report, is a voucher report, or is not either
# ERRORS:
# NOTE: Headers changed with new custom field - updated 5/9/2020

ebPOs = {}
budgTasks = {}

def checkBWcsv(theFile):
    # print "say what?"
    retVal = ""
    # OLD, pre custom fields POheader = "PO ID,Creation Date,Original Revision Date,Last Revision Date,Last Distribution Date,Workflow Completion Date,PO #,Requested Deliver Date,PR ID,Payment Type,PO Terms,Buyer: First Name,Buyer: Last Name,Buyer Phone,Buyer: Email,Department,Supplier Fax,BillTo Address Code,BillTo Address Internal Name,BillTo Contact 1,BillTo Address 1,BillTo City,BillTo State,BillTo Postal Code,BillTo Country,Accounting Date,ShipTo Address Code,ShipTo Address Internal Name,ShipTo Contact 1,ShipTo Contact 2,ShipTo Contact 3,ShipTo Address 1,ShipTo Address 2,ShipTo City,ShipTo State,ShipTo Postal Code,ShipTo Country,Supplier ID,Customer SupplierId,Supplier Name,Supplier Number,Header Notes,PO Line #,Quantity,Unit Price,Unit Price Date,Extended Price,List Price,List Price Date,Current - 1 Unit Price,Current - 1 Unit Price Date,Current - 1 List Price,Current - 1 List Price Date,SKU/Catalog #,Supplier PartAuxiliary ID,Amount/UOM & UOM,Product Size,Product Description,Shipping Method,Carrier,LineItem Notes,Non Catalog,Product Type,Supplier Duns No,Supplier Phone,Contract No,Contract Renewal No,Contract Name,Contract Effective Date,Contract Expiration Date,Contract Unit Price,Contract Unit Price Variance,FormId,Replenishment Order,Stock Item ID,Stock Item Name,Stock Units,Stock Supplier Name,Stock Supplier ID,Stock FC Name,Stock FC ID,Line Status,Manufacturer,Mfr Catalog #,Category Preference,Category Level 1,Category Level 2,Category Level 3,Category Level 4,Category Level 5,Category Name,CAS #,UNSPSC,Commodity Code,Shipping Charge,Tax 1,Tax 2,Confirming/not sent to vendor,Separate Payment,Blanket Order (Qty must = 1),Payment Hold Reason,Amount Only PO,Reject Voucher?,Payment Cancellation Date,Payment Message,AP Voucher Hold,Serial Number,Withholding Applicable,Additional Items on Invoice,PR Validation Test,Payment Term Exception,SpeedType Class,Sole Source,Requires Receiving,Withholding Class,Request Budget Override,Handling Code,Payment Term,Withholding Type,Hold Payment,Hold Vouchers in BuyWays,Activity Id,Business Unit,Project,Program,Origin Code,External Req #,FMP Number,Speedtype,Department,Account,PC Business Unit,Class,Fund,Campus PO Number Prefix,Imported PO,Price Set Name,Consortium Spend,Business Unit,Vendor Id,Shared Cart\n"
    # 200707 Apparent change in Buyways report, old one was crashing
    # 200812 Another change in Buyways report, added field "Brand" near end of report
    # POheader = "PO ID,Creation Date,Original Revision Date,Last Revision Date,Last Distribution Date,Workflow Completion Date,PO #,Requested Deliver Date,PR ID,Payment Type,PO Terms,Buyer: First Name,Buyer: Last Name,Buyer Phone,Buyer: Email,Department,Supplier Fax,BillTo Address Code,BillTo Address Internal Name,BillTo Contact 1,BillTo Address 1,BillTo City,BillTo State,BillTo Postal Code,BillTo Country,Accounting Date,ShipTo Address Code,ShipTo Address Internal Name,ShipTo Contact 1,ShipTo Contact 2,ShipTo Contact 3,ShipTo Address 1,ShipTo Address 2,ShipTo City,ShipTo State,ShipTo Postal Code,ShipTo Country,Supplier ID,Customer SupplierId,Supplier Name,Supplier Number,Header Notes,PO Line #,Quantity,Unit Price,Unit Price Date,Extended Price,List Price,List Price Date,Current - 1 Unit Price,Current - 1 Unit Price Date,Current - 1 List Price,Current - 1 List Price Date,SKU/Catalog #,Supplier PartAuxiliary ID,Amount/UOM & UOM,Product Size,Product Description,Shipping Method,Carrier,LineItem Notes,Non Catalog,Product Type,Supplier Duns No,Supplier Phone,Contract No,Contract Renewal No,Contract Name,Contract Effective Date,Contract Expiration Date,Contract Unit Price,Contract Unit Price Variance,FormId,Replenishment Order,Stock Item ID,Stock Item Name,Stock Units,Stock Supplier Name,Stock Supplier ID,Stock FC Name,Stock FC ID,Line Status,Manufacturer,Mfr Catalog #,Category Preference,Category Level 1,Category Level 2,Category Level 3,Category Level 4,Category Level 5,Category Name,CAS #,UNSPSC,Commodity Code,Shipping Charge,Tax 1,Tax 2,Confirming/not sent to vendor,Separate Payment,Blanket Order (Qty must = 1),Payment Hold Reason,Amount Only PO,Reject Voucher?,Payment Cancellation Date,Payment Message,AP Voucher Hold,Serial Number,Withholding Applicable,Additional Items on Invoice,PR Validation Test,Payment Term Exception,SpeedType Class,Sole Source,Requires Receiving,Withholding Class,Request Budget Override,Handling Code,Payment Term,Withholding Type,Hold Payment,Hold Vouchers in BuyWays,Activity Id,Business Unit,Project,Program,Origin Code,External Req #,FMP Number,Speedtype,Department,Account,PC Business Unit,Class,Fund,Campus PO Number Prefix,Imported PO,Price Set Name,Consortium Spend,Business Unit,Vendor Id,Shared Cart\n"
    POheader = "PO ID,Creation Date,Original Revision Date,Last Revision Date,Last Distribution Date,Workflow Completion Date,PO #,Requested Deliver Date,PR ID,Payment Type,PO Terms,Buyer: First Name,Buyer: Last Name,Buyer Phone,Buyer: Email,Department,Supplier Fax,BillTo Address Code,BillTo Address Internal Name,BillTo Contact 1,BillTo Address 1,BillTo City,BillTo State,BillTo Postal Code,BillTo Country,Accounting Date,ShipTo Address Code,ShipTo Address Internal Name,ShipTo Contact 1,ShipTo Contact 2,ShipTo Contact 3,ShipTo Address 1,ShipTo Address 2,ShipTo City,ShipTo State,ShipTo Postal Code,ShipTo Country,Supplier ID,Customer SupplierId,Supplier Name,Supplier Number,Header Notes,PO Line #,Quantity,Unit Price,Unit Price Date,Extended Price,List Price,List Price Date,Current - 1 Unit Price,Current - 1 Unit Price Date,Current - 1 List Price,Current - 1 List Price Date,SKU/Catalog #,Supplier PartAuxiliary ID,Amount/UOM & UOM,Product Size,Product Description,Shipping Method,Carrier,LineItem Notes,Non Catalog,Product Type,Supplier Duns No,Supplier Phone,Contract No,Contract Renewal No,Contract Name,Contract Effective Date,Contract Expiration Date,Contract Unit Price,Contract Unit Price Variance,FormId,Replenishment Order,Stock Item ID,Stock Item Name,Stock Units,Stock Supplier Name,Stock Supplier ID,Stock FC Name,Stock FC ID,Line Status,Manufacturer,Mfr Catalog #,Category Preference,Category Level 1,Category Level 2,Category Level 3,Category Level 4,Category Level 5,Category Name,CAS #,UNSPSC,Commodity Code,Shipping Charge,Tax 1,Tax 2,Confirming/not sent to vendor,Separate Payment,Blanket Order (Qty must = 1),Payment Hold Reason,Amount Only PO,Reject Voucher?,Payment Cancellation Date,Payment Message,AP Voucher Hold,Serial Number,Withholding Applicable,Additional Items on Invoice,PR Validation Test,Payment Term Exception,SpeedType Class,Sole Source,Requires Receiving,Withholding Class,Request Budget Override,Handling Code,Payment Term,Withholding Type,Hold Payment,Hold Vouchers in BuyWays,Activity Id,Business Unit,Project,Program,Origin Code,External Req #,FMP Number,Speedtype,Department,Account,PC Business Unit,Class,Fund,Campus PO Number Prefix,Imported PO,Price Set Name,Consortium Spend,Business Unit,Vendor Id,Shared Cart,Brand,Original Requisition ID,Original Requisition Name,Original Requisition Requestor\n"
    # VoucherHeader = "Workflow Status,Invoice No,Voucher Type,Invoice Date,Invoice Status,PO No,PO Line No,Workflow Status,Invoice Line Extended Price,Account,Activity Id,Additional Items on Invoice,Amount Only PO,AP Voucher Hold,Blanket Order (Qty must = 1),Business Unit,Campus PO Number Prefix,Class,Confirming/not sent to vendor,Department,External Req #,Fund,Handling Code,Hold Payment,Hold Vouchers in BuyWays,Imported PO,Origin Code,Payment Cancellation Date,Payment Hold Reason,Payment Message,Payment Term,Payment Term Exception,PC Business Unit,PR Validation Test,Program,Project,Reject Voucher?,Request Budget Override,Requires Receiving,Separate Payment,Serial Number,Sole Source,Speedtype,SpeedType Class,Withholding Applicable,Withholding Class,Withholding Type,Serial Numbers\n"
    VoucherHeader = "Workflow Status,Invoice No,Supplier Name,Supplier Number,Supplier Invoice No,PO No,Invoice System Created Date,Invoice System Created Time,Invoice Date,Invoice Due Date,Invoice Discount Date,Voucher Source,Voucher Type,Invoice Total,Invoice Status,PO Line No,Catalog No,Product Name,Invoice Line Unit Price,Invoice Line Currency,Quantity,UOM,Invoice Line Extended Price,Department Name,Contract No,Contract Renewal No,Contract Name,Contract Effective Date,Contract Expiration Date,Contract Unit Price,Contract Unit Price Variance,Previously Sent to ERP,Export Date,Workflow Complete Date,Apply Early Payment Discount,Early Payment Discount,PO Business Unit,PO Business Unit Vendor Id,Invoice Discount,Tax1 Charge,Tax2 Charge,Shipping Charge,Handling Charge,Paid Date,Payable Date,Accounting Date,Account,Activity Id,Additional Items on Invoice,Amount Only PO,AP Voucher Hold,Blanket Order (Qty must = 1),Business Unit,Campus PO Number Prefix,Class,Confirming/not sent to vendor,Department,External Req #,FMP Number,Fund,Handling Code,Hold Payment,Hold Vouchers in BuyWays,Imported PO,Origin Code,Payment Cancellation Date,Payment Hold Reason,Payment Message,Payment Term,Payment Term Exception,PC Business Unit,PR Validation Test,Program,Project,Reject Voucher?,Request Budget Override,Requires Receiving,Separate Payment,Serial Number,Sole Source,Speedtype,SpeedType Class,Withholding Applicable,Withholding Class,Withholding Type,Serial Numbers\n"
    try:
        fp = open(theFile)
        last_pos = fp.tell()
        line = fp.readline()
        # 200831
        """
        if line == POheader:
            retVal = "isPO"
        elif line == VoucherHeader:
            retVal = "isVoucher"
        else:
            retVal = "notBW"
        """
        if line[:40] == POheader[:40]:
            retVal = "isPO"
        elif line[:40] == VoucherHeader[:40]:
            retVal = "isVoucher"
        else:
            retVal = "notBW"

        fp.seek(last_pos)
        line = fp.readline()
        fp.seek(last_pos)
        fp.close()
    except:
        retVal = "File Not Found"
    return retVal
def parse_Buyways_Description(BWdesc):
    # 230727 We need to expand this logic: if the FMP is not in the PS project number, we need to look in the description of the first line
    # This is where Bus Ops is adding it as "FMP XXXXXX"
    retVal = "NON FMP"
    toks = BWdesc.split("FMP")
    if len(toks) > 1:
        theFMP = toks[1][1:7]
        if theFMP.isnumeric():
            retVal = theFMP
        else:
            print("Error in parse_Buyways_Description - not 6 digit number")
    # print(">>>*&*&", retVal)
    return retVal
            
def parse_Buyways_Project(proj):
    #CAUTION: buyways allows multiple numbers separared by |
    #AND, WAIT FOR IT !!!! "FMP" entered inconsistently, not always chars 9-15
    # Need to search for it? Or??
    # EXAMPLES:
    #    M50190FMP319020-L
    #   M5019FMP0317381-L
    # MORE EXAMPLES/CONDITIONS
    # Logic: Find "FMP" (can be multiple", FInd "-L". If delta of those is 10, the
    # the FMP has leading 0, if 9 it doesn't. Multiple project numbers are sep.
    # by a "|". FMP can be first, last, middle, all... Delta can be negative
    #    none|M5019FMP0315143-L
    #   none|M5013FMP0224798-L
    # M50150000000001-L|M5019FMP0301042-L
    # M50080000000005-L|M50110000000002-L|M5018FMP0301045-L
    # Maybe split by "|", then look at each...
    #What about M5019FMP0301105-L? 301105 is EXACTLY what we're trying to find:
    # an EB Active Project, with budget, that Buyways has a PO for but EB doesn't
    i = proj.find("FMP")
    j = proj.find("-")
    if i>0: # confirm it's an FMP number
        FMP = proj[j-6:j] # why hard coded? It could have a 0
        #if FMP == "319020" or FMP == "301105":
            #print ">>>>>",FMP,"<<<<<<",len(FMP)
        #print "****", FMP
    else:
        FMP = "NON FMP"
        #if FMP == "NON FMP" then parse HeaderNotes to look for FMP
    return FMP

def parseHeaderNotes(headerNotes):
    # Find the index of 'FMP' in the input string
    fmp_index = headerNotes.find('FMP')

    # If 'FMP' is found, extract the characters after it
    if fmp_index != -1 and len(headerNotes) > fmp_index + 3:
        fmp_number = headerNotes[fmp_index + 3:].lstrip('0')
        if fmp_number:  # Check if there are any non-zero characters
            return fmp_number
    return "NON FMP"

def getBudgetTasks():
    # print("Defining budget lines with tasks")
    global budgetTasks
    budgetTasks = {"772130":"20.772130",
                    "772140":"40.772140",
                    "772160":"30.772160",
                    "772180":"60.772180",
                    "772220":"30.772220",
                    "772230":"30.772230",
                    "734200":"60.734200",
                    "734310":"10.734310",
                    "734800":"50.734800",
                    "735100":"10.735100",
                    "735200":"80.735200",
                    "735600":"10.735600",
                    "740800":"60.740800",
                    "741970":"60.741970",
                    "741980":"50.741980",
                    "744300":"60.744300",
                    "744700":"30.744700",
                    "757060":"20.757060",
                    "757120":"20.757120",
                    "757190":"10.757190",
                    "757220":"70.757220",
                    "761080":"50.761080",
                    "761090":"60.761090",
                    "761210":"40.761210",
                    "761211":"40.761211",
                    "761460":"10.761460",
                    "761590":"70.761590",
                    "761600":"30.761600",
                    "761603":"30.761603",
                    "761604":"30.761604",
                    "763100":"50.763100",
                    "763120":"50.763120",
                    "763140":"50.763140",
                    "763210":"60.763210",
                    "763400":"30.763400",
                    "763900":"60.763900",
                    "764300":"50.764300",
                    "766100":"50.766100",
                    "766300":"60.766300",
                    "766600":"60.766600",
                    "767210":"60.767210",
                    "767300":"60.767300",
                    "767800":"60.767800",
                    "768100":"50.768100",
                    "768210":"60.768210",
                    "768300":"60.768300",
                    "768600":"60.768600",
                    "772010":"20.772010",
                    "772050":"10.772050",
                    "772060":"20.772060",
                    "772120":"10.772120",
                    "772130":"20.772130",
                    "772140":"40.772140",
                    "772160":"30.772160",
                    "772180":"60.772180",
                    "772220":"30.772220",
                    "772230":"30.772230",
                    "700T10":"10.700T10",
                    "700T20":"20.700T20",
                    "700T30":"30.700T30",
                    "700T40":"40.700T40",
                    "700T50":"50.700T50",
                    "700T60":"60.700T60",
                    "700T70":"70.700T70",
                    "700T80":"80.700T80",
                    "99CONT":"99.99CONT"}
    return budgetTasks

POprocessHeaders = {1:"FMP Number",2:"Process Counter",3:"Commitment Number",4:"PeopleSoft PO#",5:"PeopleSoft PO Amount",\
                        6:"Step",7:"Speedtype",8:"Origin Code",9:"Notes"}

# These are the headers for the excel POCostImportJoinedData
POcostHeaders = {1:"Commitment Number",2:"FMP Number",3:"Commitment Type",4:"Company Name",5:"Budget Line Item",6:"Status",7:"Item Number",\
                     8:"Description",9:"Item Quantity",10:"Item Unit Cost",11:"Commitment Date",12:"Company Number",13:"Commitment Item Amount",\
                     14:"Item Unit of Measure",15:"Funding Rule",16:"Speedtype",17:"PeopleSoft PO#",18:"Origin Code",19:"Commodity Code",20:"Retainage",21:"Notes"}

InvoiceProcessHeaders = {1:"FMP Number",2:"Process Counter",3:"Step",4:"Status",5:"Commitment Number",6:"Vendor Invoice #",7:"Voucher ID"}

InvoiceCostHeaders = {1:"FMP Number",2:"Invoice Number",3:"Vendor Invoice #",4:"Commitment Number",5:"Description",6:"Invoice Item Amount",\
                          7:"Commitment Item Number",8:"Funding Rule",9:"Voucher ID",10:"Status",11:"Approved Date",\
                          12:"Amount this Period",13:"Stored Materials",14:"Quantity This Period"}
updateInvoiceHeaders = {1:"FMP Number",2:"Status",3:"Invoice Number",4:"Company Name",5:"Company Number",6:"Date Paid"}

fundingRules = {"17039":"UMass Facilities","17040":"Commonwealth Facility Improvements","17101":"DCAMM Bond Funding","51161":"General Operating Fund",\
                    "51425":"Internal UML Projects","53206":"State Direct G&C Level 6","57000":"Campus Renovations","57326":"UMBA Bond Funds"}

def write_ExcelHeaders(outWB, outWS, ebType):
    # these may be better set elsewhere: they MUST be coordinated with the data fields for each row
    #print "Write Excel Headers", ebType
    if ebType == "POprocess":
        currHeaders = POprocessHeaders
    elif ebType == "POcost":
        currHeaders = POcostHeaders
    elif ebType == "InvoiceProcess":
        #print "OK - InvoiceProcess"
        currHeaders = InvoiceProcessHeaders
    elif ebType == "InvoiceCost":
        currHeaders = InvoiceCostHeaders
    elif ebType == "updateInvoice":
        currHeaders = updateInvoiceHeaders

    for h in range(1,len(currHeaders)+1):
        outCell = outWS.cell(row=1,column=h)
        outCell.value = currHeaders[h]


def create_Excel():
    wbOut = Workbook()
    #sheetName = "EBimport"
    #outWS = wbOut.create_sheet(sheetName,0)
    # use default, "Sheet" - CONSIDER renaming to "EBimport"
    outWS = wbOut.active
    return(wbOut,outWS)

def parseST(ST):
    # Speedtypes in Buyways are separated by "|"
    # and they haaave a "-L" appended, that we don't want in EB
    #print ST
    toks = ST.split('-L|')
    # The last Speedtype in the list will still have "-L"
    # so strip it away
    if len(toks) != 1:
        toks[len(toks)-1] = toks[len(toks)-1][:-2]
    #print toks, len(toks)
    return toks

def parseStr(s):
    safeStr = ""
    for c in s:
        if c in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789 /-.&$":
            safeStr += c
        else:
            safeStr += " "
    #print safeStr
    if len(safeStr) > 400:
        safeStr = safeStr[:400]
    return safeStr

### NOT CHECKED FOR NEW API 1/31/2023


# PURPOSE:  Build dictionary of POs in Active EBuilder projects, with funding rules, speedtypes, FMP numbers, (?and PeopleSoft Project ID?)
# METHOD:   Read standard report from E-Builder
# INPUT:    Excel file exported from E-Builder report "PGB Buyways Import Support", saved to X: drive folder
# OUTPUT:   {PO:[funding rules list],[speedtype list],FMP,
# ERRORS:

def ebSupportData(supportFile):
    # 201202
    # Not using fixed location for support report Excel
    # Now, specififed by user
    # REVISED SUBSTANTIALLY 5/18/2020
    # Speedtype, FMP, Funding Source Category Name
    # EB Report: PGB Import Support v2
    # "I:\\Projects\\ebuilder\\reconcile\\Buyways Import Support CURRENT.xlsx"
    # supportFile = "X:\\Workspace\\Transfer\\ebReconcile\\Buyways Import Support TEST.xlsx"
    # supportFile =  "I:\\Projects\\ebuilder\\reconcile\\Buyways Import Support CURRENT.xlsx"
    supportWB = load_workbook(filename=supportFile)
    supportWS = supportWB["Sheet1"]

    ebST = {}

    print("Checking support report")
    i = 0
    for row in supportWS.iter_rows():
        if i > 0:
            currST = str(row[0].value)
            currFMP = str(row[1].value)
            currFSCN = str(row[2].value)
            # print currST, currFMP, currFSCN
            if currST in ebST:
                print("Duplicate:", currST, currFMP, currFSCN, i)
            else:
                ebST[currST] = {"FMP": currFMP, "FundSource": currFSCN}

            # d[currPO] = {"FMP":currFMP,"FundRules":[currFundRule],"Speedtypes":[currSpeedtype],"PSpo":currPSpo}
        i += 1

    # for p in ebPOs:
    #   print p, ebPOs[p]["FundRules"], ebPOs[p]["Speedtypes"]
    return ebST


def lookupPO(p):
    # Old POs, that were imported to Cost directly, are named by PeopleSoft PO number, like L000787878. Newer are named
    # by Process Instance name, like POREQ - 00023. In both cases, the PO has a custom field with the PS PO number in it

    # These are from the EB report ---
    global ebPOs
    # We need to add POs from API
    # ebAPI_POs =
    retVal = "NOT IN EB"
    # print p

    if p in ebPOs:  # it's an old PO, with PS PO number as commitment name
        # print "Found as committment name"
        retVal = p
    else:
        for q in ebPOs:
            # print "Checking custom field", p, q, ebPOs[q]["PSpo"]
            if retVal != "NOT IN EB":
                break
            if ebPOs[q]["PSpo"] == p:
                retVal = q
    # print ">>>>", retVal
    return retVal


def lookupSpeedtype(s):
    STinEB = False
    global ebPOs

    for q in ebPOs:
        for st in ebPOs[q]["Speedtypes"]:
            if STinEB == True:
                break
            else:
                # print st,s
                # print len(st), len(s)
                if st == s[:-2]:
                    # print "!@!!!!!!!!!!!!!!!!!!! Found speedtype", st, ebPOs[q]["FMP"]
                    STinEB = True
                    currST = st
                    currFMP = ebPOs[q]["FMP"]
                    currFundRule = ebPOs[q]["FundRules"][0]
                # else:
                # print "WT?", st, s, id(st), id(s)

    if STinEB == False:
        # print "Did not find speedtype", s
        return "NO ST", "NON FMP", "NO FUND RULE"
    else:
        return currST, currFMP, currFundRule


def buildFundingRule(currST, fundID):
    # input will be from BW report. We are going on faith that Budget Office
    # correctly named funding Rule
    # CAREFUL: Not sure how often these get updated!!!!!!!!!!
    retVal = ""

    funds = {"57000": "Campus Renovations",
             "17039": " UMass Facilities",
             "51425": " Internal UML Projects",
             "17050": " Small Repairs Projects",
             "51511": " Core Research Facilities",
             "53206": " State Direct G&C Level 6",
             "17035": " Demand Response",
             "51199": " Resident Halls",
             "51161": " General Operating Fund",
             "57326": " UMBA Bond Funds",
             "51402": "  Cost Share Match UML",
             "50000": " DCAMM"}

    try:
        retVal = funds[fundID] + "-" + currST
    except:
        retVal = "CHECK FUNDING RULE"

    return retVal

def joinReports(files):
    all_data = []
    # for index, filename in enumerate(files):
    for index,filename in enumerate(files[1:]):
        if filename.endswith('.xlsx'):
            # Load the Excel file
            workbook = load_workbook(filename)
            worksheet = workbook.active
            data = []
            for row in worksheet.iter_rows(values_only=True):
                data.append(row)
            if index != 0:
                data = data[1:]
            all_data.extend(data[1:])
            # Delete the file
            os.remove(filename)

    # Create a new workbook for the joined data
    joined_workbook = load_workbook(files[0])
    os.remove(files[0])
    joined_worksheet = joined_workbook.active
    for row in all_data:
            joined_worksheet.append(row)
    # Apply left alignment to all cells
    for row in joined_worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left')

    return joined_workbook

def makeStatsHTML(stats,title):
    htmlHead = "<!DOCTYPE html>\n<html>\n<head>\n<title>"+title+"</title>\n</head>\n"
    htmlStyle = "<style>\n"
    htmlStyle += "table {font-family: Open Sans, sans-serif; border-collapse: collapse; width: 100%; margin-bottom: 20px;}\n"
    htmlStyle += "td, th {border: 1px solid #ddd; padding: 8px; text-align: left;}\n"
    htmlStyle += "th {background-color: #0463a7; color: white;}\n"
    htmlStyle += "tr:nth-child(even) {background-color: #f9f9f9;}\n"
    htmlStyle += "body{ font-family: 'Open Sans',sans-serif;}\n"
    htmlStyle += "</style>\n"

    totalsDF = pd.DataFrame(stats, index=[0])
    tableData = totalsDF.to_html(index=False)

    htmlBody = "<body>\n<h2>"
    htmlBody += title + "</h2>\n"
    htmlBody += tableData
    htmlBody += "\n</body>\n</html>"

    htmlData = htmlHead + htmlStyle + htmlBody
    return htmlData


    
def removeXMLFile(f, lineType):
    #240104 We should build a list and then delete items from it - faster?
    print("***** what?")
    try:
        print("***** deleting:", f)
        os.remove(f)
        print(f, "belongs to a ", lineType, "PO. So, Deleting" )
        # print(f,' is not a UML PO so removing')
    except Exception as e:
        print("Error:", str(e))
        ignoreThis = 1
        print("File doesn't exist or deleted in previous invoice line")


# get_BudgetTasks
# PURPOSE:  Define budget line item IDs with associated "Tasks"
# METHOD:   Hard coded list from e-builder, which originated in PeopleSoft
# INPUT:    None
# OUTPUT:   Dictionary of line item: task.line item
# ERRORS:
# NOTES:    EB does not have all budget lines. Can create problems when BW uses budget line that's not
#           in EB


pyFile = os.path.basename(__file__)
print(pyFile, "reloaded")
