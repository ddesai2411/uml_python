#!/usr/bin/env python
# coding: utf-8
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime

#import newebapi_proj as proj_data
#import newebapi_budg as budg_data

import uml_python.uml_lib.ebAPI_lib as ebAPI

def newliner(s):
    # replace eb output of /n/n with Excel-ready \n\n.
    # Watch out for /'s used as part of text, like "Design/Build"
    retVal = ""
    # strip /n or /n/n off the end of the string
    if s[-4:] == "/n/n":
        s = s[:(len(s)-4)]
    elif s[-2:] == "/n":
        # Need to add connection to Budgets module, to get TPC cost cust
        s = s[:(len(s)-2)]

    skipCount = 0
    for i in range(0,len(s)):
        if skipCount == 0:
            if s[i] == '/':
                #print ("!!!!!!!!!",i, len(s), s[i:i+4])
                if i < (len(s)-2) and s[i:i+2] == "/n":
                    retVal += "\n"
                else:
                    retVal += s[i:i+2]
                skipCount = 1
            else:
                retVal += s[i]
        else:
            skipCount -= 1
    #print retVal
    return retVal

def write_cell(ws, r,c,val):
    outCell = ws.cell(row=r, column=c)
    outCell.value = val

"""
def convertDate(ut):
    # unix time to day, month, year string
    # ut is in milliseconds
    ut = ut[6:19]
    ut = int(ut)
    #timeStr = datetime.utcfromtimestamp(ut/1000).strftime('%Y-%m-%dT%H:%M:%SZ')
    timeStr = datetime.utcfromtimestamp(ut/1000).strftime('%m.%d.%Y')
    monthStr = datetime.utcfromtimestamp(ut/1000).strftime('%m %Y')
    return timeStr
"""
# Format of the date is changed in New API

def convertDate(ut):
    timeStr = datetime.fromisoformat(ut).strftime('%m.%d.%y')
    return timeStr

def get_monthYear():
    retVal = ""
    months = ["Unknown","January","Febuary","March","April","May","June",
          "July","August","September","October","November","December"]
    now = datetime.now()
    if now.month == 12:
        mo = "January"
        retVal = mo + " " + str(now.year + 1)
    else:
        mo = months[now.month+1]
        retVal = mo + " " + str(now.year)
    return(retVal)

def get_numeric_monthYear():
    months = ["Unknown","January","Febuary","March","April","May","June",
          "July","August","September","October","November","December"]
    now = datetime.now()
    #moStr = str(now.month-4)
    if now.month == 12:
        moStr = "01"
        yrStr = str(now.year+1)[2:]
    else:
        mo = months[now.month+1]
        moStr = str(mo)
        if len(moStr) < 2:
            moStr = "0" + moStr
        yrStr = str(now.year)[2:]
    #print moStr, yrStr
    return (yrStr+ moStr + "_")

def get_TPC(budget_data, currProjID):
    currTPC = -1
    a = len(budget_data)

    for b in range(0,a):
        if budget_data[b]["projectId"] == currProjID:
            #print ">>>>", budget_data[b]["ProjectID"], currProjID
            #print "!!!!!!",budget_data[b]["Custom_Total_Current_Working_EstimateTPC"]
            try:
                currTPC = budget_data[b]["Total Current Working Estimate/TPC"]
            except:
                currTPC = -1
            break

    if currTPC == -1:
        currTPC = None
    #print ("RETURNING ............................................... TPC", currTPC)
    return currTPC

def checkLead(dept, PM, Planner):
    retVal = True
    # we don't want to output if the lead department's person is TBD or blank
    if dept == "Project Management":
        if PM == "TBD" or PM == "":
            retVal = False
    elif dept == "Planning":
        if Planner == "TBD" or Planner == None:
            retVal = False
    else:
        "Dept is not PM or Planning", dept
    return retVal

def getDataFromAPI(processPrefix):
    datafields = {
        "selectedfields": [
            "Project/ProjectName",
            "Project/CustomFields/FMP Number",
            "Project/Status",
            "Project/CustomFields/FM Department Lead",
            "Project/CustomFields/Status Comments",
            "ProcessInstance/CurrentStepName"
        ],
        "Filters": [
            {
                "Field": "ProcessInstance/CurrentStepName",
                "Operation": "=",
                "Value": "FIS Review"
            },
        ],
    }

    theURL = 'https://api2.e-builder.net/api/v2/noncostprocesses/query?processprefix=' + str(processPrefix)
    print('Getting ' + str(processPrefix) +' data')
    data = ebAPI.postTOAPI(theURL, datafields)['records']
    return data

def getCloseoutFMPs():
    closeoutFMPs = []
    CLOjson = getDataFromAPI("CLO")
    CLNONjson = getDataFromAPI("CLNON")
    for i in range(len(CLOjson)):
        FMP = CLOjson[i]['Project']['CustomFields']['FMP Number']
        if FMP not in closeoutFMPs:
            closeoutFMPs.append(FMP)
    for i in range(len(CLNONjson)):
        FMP = CLNONjson[i]['Project']['CustomFields']['FMP Number']
        if FMP not in closeoutFMPs:
            closeoutFMPs.append(FMP)
    return closeoutFMPs



def main():

    # Monthly Status Reports(Month of Report)
    retVal = {"Month":"-","ProjCount":0,"No Updates": 0, "Not Active":0, "ofile":"","obj length":0}
    # 0> Create Excel
    outXL  = Workbook()
    outXL_WS = outXL.active
    moStatHeaders = {1:"FMP Number",2:"Previous Status Comments",3:"FM Department Lead",4:"Project Health",\
                     5:"Project Phase", 6: "Successor Projects?",7:"Enabling Projects", 8: "Current Project Start Date",\
                     9:"Current Project End Date",10:"Step",11:"Total Current Working Estimate/TPC",12:"Monthly Status Report (Month of Report)",
                     13:"Project Manager",14:"Planner"}
    # NOTE: Other fields to add/review:
    # Project Health, Project Phase, Successor Projects?, Enabling Projects?
    # Current Project Start Date
    # Current Project End Date

    monthYear = get_monthYear() # assumes we're running this program for reports for the following month. EG: we run it in October for November reports
    #monthYear = "January 2023"
    retVal["Month"] = monthYear
    #retVal["Month"]

    for h in range(1,len(moStatHeaders)+1):
        outCell = outXL_WS.cell(row=1,column=h)
        outCell.value = moStatHeaders[h]

    # Make previous status wide for legibilty in Excel
    outXL_WS.column_dimensions['B'].width = 60.

    #project_data = ebAPI.get_active_project_all_data()
    project_data = ebAPI.getDataFromCache("ActiveProjects")
    print('Active Projects data imported')
    #budget_data = ebAPI.get_budget_all_data()
    budget_data = ebAPI.getDataFromCache("Budgets")
    print('Budgets data imported')

    closeoutFMPs = getCloseoutFMPs()

    i = len(project_data)
    retVal["obj length"] = i
    r = 2 #counter for row

    for j in range(0,i):
        #try:
            # strip trailing white space?
            # 201203 No exceptions/crashes with these fields
            currProjName = project_data[j]["name"]#.encode("utf-8")
            currFMP = project_data[j]["FMP Number"]#.encode("utf-8")
            currStatus = project_data[j]["status"] # Project Status
            #print ">>>>", currProjName, currFMP, currStatus

            currLeadDept = project_data[j]["FM Department Lead"]
            currStatusComments = project_data[j]["Status Comments"] #newliner here?
            currProjID = project_data[j]["projectId"] # Need this for lookup in Budget Data
            currPM = project_data[j]["Project Manager"]
            currPlanner = project_data[j]["Project Planner"]
            currPhase = project_data[j]["Project Phase"]
            if currLeadDept == "Planning" or currLeadDept == "Project Management":
                reportable = True # FIS or O&S, don't report
                definedLead = checkLead(currLeadDept, currPM, currPlanner)
            else:
                reportable = False
            if currFMP in closeoutFMPs:
                closeout = True
            else:
                closeout = False
            if reportable == True and closeout == False:
                if currStatus == "Active" or currStatus == "TD Active":
                    print (">>>>", currPhase, currPhase[:2])
                    if currPhase[:2] != "08":
                        #print "Not 08", currPhase
                        if definedLead == True:
                            #print currFMP
                            currTPC = get_TPC(budget_data, currProjID)

                            if currLeadDept == "Project Management":
                                currStep = "PM Review"
                            elif currLeadDept == "Planning":
                                currStep = "Planning Review"
                            else:
                                currStep = "IGNORE: no updates"
                            #print currStatus, currTPC,currLeadDept,currStep
                            if currStep != "IGNORE: no updates":
                                write_cell(outXL_WS,r,1,currFMP)
                                try:
                                    currComments_w_newlines = newliner(currStatusComments)
                                except:
                                    currComments_w_newlines = "No previous comments"
                                #print "\n?????\n", currComments_w_newlines
                                write_cell(outXL_WS, r,2,currComments_w_newlines)
                                cellStr = "B" + str(r)
                                outXL_WS[cellStr].alignment = Alignment(wrapText=True)
                                write_cell(outXL_WS,r,3,currLeadDept)

                                if project_data[j]["Project Health"] == None:
                                    currProjHealth = "TBD"
                                else:
                                    currProjHealth = project_data[j]["Project Health"]
                                write_cell(outXL_WS,r,4,currProjHealth)#.encode("utf-8"))

                                write_cell(outXL_WS,r,5,project_data[j]["Project Phase"])#.encode("utf-8"))
                                try:
                                    currSuccessorProjects = project_data[j]["Successor Projects"]#.encode("utf-8")
                                except:
                                    currSuccessorProjects = ""
                                write_cell(outXL_WS,r,6,currSuccessorProjects)
                                try:
                                    currEnablingProjects = project_data[j]["Enabling Projects"]#.encode("utf-8")
                                except:
                                    currEnablingProjects = ""
                                write_cell(outXL_WS,r,7,currEnablingProjects)

                                # Date formats: need to confirm what EB wants and figure out what Excel does to it
                                # may require quoting it as a string or formatting it
                                # Raw text from API/web version is <d:StartDate m:type="Edm.DateTime">2019-09-17T00:00:00</d:StartDate>
                                timeStr = convertDate(project_data[j]["startDate"])
                                write_cell(outXL_WS,r,8,timeStr)
                                timeStr = convertDate(project_data[j]["targetDate"])
                                write_cell(outXL_WS,r,9,timeStr)
                                write_cell(outXL_WS,r,10,currStep)
                                write_cell(outXL_WS,r,11,currTPC)
                                write_cell(outXL_WS,r,12,monthYear)
                                write_cell(outXL_WS,r,13,currPM)
                                write_cell(outXL_WS,r,14,currPlanner)
                                r += 1
                            else:
                                #print "Step was no updates?", currFMP, currProjName
                                retVal["No Updates"] += 1
                    else:
                        #print "Lead is not defined (TBD or None)"
                        ignoreThis = 1
                        print ("Undefined lead", currFMP, currProjName, currLeadDept, currPM, currPlanner)
                else:
                    #print "Ignoring - not Active or TD Active", currFMP, currProjName
                    retVal["Not Active"] += 1

    retVal["ProjCount"] = r-1
    monthStr = get_numeric_monthYear()
    #uname = getpass.getuser()
    # oDir = "/Users/kysgattu/FIS/uml_python/uml/outputfiles/"
    oDir = "B:\\MonthlyStatusReports\\"
    ofile = oDir + monthStr + "monthlyStatus.xlsx"
    retVal["ofile"] = ofile
    print(ofile)
    outXL.save(ofile)
    return retVal

if __name__ == "__main__":
    main()

#get_ipython().system('jupyter nbconvert --to script moStat_v1.ipynb')

